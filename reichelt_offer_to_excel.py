#!/usr/bin/env python3
"""
Convert a Reichelt offer PDF into the Excel format expected by ariba_excel_import.py.

Mapping:
- Reichelt "Item No."    -> Excel "Product name"
- Reichelt "Description" -> Excel "Description"
- Quantity               -> Quantity
- Price                  -> Unit price

Ignored columns:
- Category of goods
- Price on all
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook


@dataclass
class OfferItem:
    item_no: str
    description: str
    quantity: str
    unit_price: str


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_euro_number(text: str) -> str:
    value = text.strip()
    value = re.sub(r"[^0-9,.\-]", "", value)
    value = value.replace(",", ".")
    return value


def extract_text_with_pypdf(pdf_path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return ""

    reader = PdfReader(str(pdf_path))
    chunks: List[str] = []
    for page in reader.pages:
        # pypdf supports layout extraction in recent versions; fallback to default.
        try:
            text = page.extract_text(extraction_mode="layout")
        except TypeError:
            text = page.extract_text()
        chunks.append(text or "")
    return "\n".join(chunks)


def extract_text_with_pdftotext(pdf_path: Path) -> str:
    tool = shutil.which("pdftotext")
    if not tool:
        return ""
    try:
        result = subprocess.run(
            [tool, "-layout", "-enc", "UTF-8", str(pdf_path), "-"],
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout
    except Exception:
        return ""


def iter_table_lines(full_text: str) -> Iterable[str]:
    in_table = False
    for raw in full_text.splitlines():
        line = raw.rstrip()
        normalized = line.lower()

        if "item no." in normalized and "description" in normalized and "quantity" in normalized:
            in_table = True
            continue

        if not in_table:
            continue

        if "order value" in normalized:
            break

        if not line.strip():
            continue

        # Skip wrapped header fragments and non-row table text.
        if normalized.strip() in {"of goods", "incl. vat"}:
            continue

        yield line


def parse_table_lines(lines: Iterable[str]) -> List[OfferItem]:
    items: List[OfferItem] = []

    for line in lines:
        parts = re.split(r"\s{2,}", line.strip())
        if len(parts) < 6:
            continue

        item_no = clean_spaces(parts[0])
        description = clean_spaces(parts[1])
        category = clean_spaces(parts[2])
        quantity = clean_spaces(parts[3])
        price = clean_spaces(parts[4])

        if not category.isdigit():
            continue
        if not quantity.isdigit():
            continue

        unit_price = parse_euro_number(price)
        if not re.fullmatch(r"\d+(?:\.\d+)?", unit_price):
            continue

        if not item_no or not description:
            continue

        items.append(
            OfferItem(
                item_no=item_no,
                description=description,
                quantity=quantity,
                unit_price=unit_price,
            )
        )

    return items


def write_output_excel(output_path: Path, supplier_name: str, items: List[OfferItem]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["B1"] = supplier_name

    ws["A3"] = "Product name"
    ws["B3"] = "Description"
    ws["C3"] = "Quantity"
    ws["D3"] = "Unit price"

    row = 4
    for item in items:
        # Mapping:
        # Reichelt Item No.    -> Product name
        # Reichelt Description -> Description
        ws.cell(row=row, column=1, value=item.item_no)
        ws.cell(row=row, column=2, value=item.description)
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=item.unit_price)
        row += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert a Reichelt offer PDF into Ariba import Excel format."
    )
    parser.add_argument("--pdf", required=True, help="Path to Reichelt offer PDF.")
    parser.add_argument(
        "--out",
        required=True,
        help="Path to output .xlsx file (for ariba_excel_import.py).",
    )
    parser.add_argument(
        "--supplier",
        default="Reichelt",
        help="Supplier name written to cell B1 (default: Reichelt).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pdf_path = Path(args.pdf).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()

    if not pdf_path.exists():
        print(f"Input PDF not found: {pdf_path}")
        return 2

    text = extract_text_with_pypdf(pdf_path)
    if not text.strip():
        text = extract_text_with_pdftotext(pdf_path)

    if not text.strip():
        print(
            "Could not extract text from PDF. Install pypdf or ensure pdftotext is available."
        )
        return 2

    items = parse_table_lines(iter_table_lines(text))
    if not items:
        print("No offer rows found in PDF table.")
        return 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(out_path, args.supplier, items)

    print(f"Wrote {len(items)} item(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
