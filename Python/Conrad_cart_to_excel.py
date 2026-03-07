#!/usr/bin/env python3
"""
Convert a Conrad cart CSV into the Excel format expected by ariba_excel_import.py.

Mapping:
- Conrad "Conrad Article-Nr. | Description" (max 80 chars) -> Excel "Product name"
- Conrad "Description" -> Excel "Description"
- Conrad "Quantity" -> Excel "Quantity"
- Conrad "Unit Price" (gross) / (1 + VATRate.xlsx[A2]/100) -> Excel "Unit price"
- Conrad "Conrad Article-Nr." -> Excel "Supplier Part Number"
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_decimal_text(value: object, field_name: str) -> str:
    text = clean_text(value)
    text = re.sub(r"[^0-9,.\-]", "", text).replace(",", ".")
    if not text:
        raise ValueError(f"Missing value for '{field_name}'.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {field_name} value '{value}'.") from exc
    if number <= 0:
        raise ValueError(f"{field_name} must be > 0 (got '{value}').")
    return format(number.normalize(), "f")


def find_column(fieldnames: List[str], exact_name: str, prefix_name: str) -> str:
    if exact_name in fieldnames:
        return exact_name
    for name in fieldnames:
        if name.startswith(prefix_name):
            return name
    raise ValueError(f"CSV missing required column: '{exact_name}'")


def build_product_name(article_number: str, description: str, max_len: int = 80) -> str:
    combined = f"{article_number} | {description}"
    return combined[:max_len]


def load_vat_rate_percent(vat_xlsx_path: Path) -> Decimal:
    try:
        workbook = load_workbook(vat_xlsx_path, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Unable to read VAT rate file: {vat_xlsx_path}") from exc

    raw_value = workbook.active["A2"].value
    if raw_value is None:
        raise RuntimeError(f"VAT rate cell A2 is empty in: {vat_xlsx_path}")

    normalized = str(raw_value).strip().replace("%", "").replace(",", ".")
    try:
        vat_rate = Decimal(normalized)
    except InvalidOperation as exc:
        raise RuntimeError(f"VAT rate in A2 is not a valid number: {raw_value!r}") from exc

    if vat_rate < 0:
        raise RuntimeError(f"VAT rate in A2 must be >= 0, got: {vat_rate}")
    return vat_rate


def gross_to_net_price(gross_price: str, vat_rate_percent: Decimal) -> str:
    gross = Decimal(gross_price)
    divisor = Decimal("1") + (vat_rate_percent / Decimal("100"))
    net = (gross / divisor).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    net_text = format(net, "f").rstrip("0").rstrip(".")
    return net_text or "0"


def parse_conrad_csv(csv_path: Path, vat_rate_percent: Decimal) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames or []
        if not fieldnames:
            raise ValueError("CSV has no header row.")

        quantity_col = find_column(fieldnames, "Quantity", "Quantity")
        article_col = find_column(fieldnames, "Conrad Article-Nr.", "Conrad Article-Nr.")
        description_col = find_column(fieldnames, "Description", "Description")
        unit_price_col = find_column(fieldnames, "Unit Price", "Unit Price")

        for row in reader:
            article_number = clean_text(row.get(article_col, ""))
            description = clean_text(row.get(description_col, ""))
            quantity_raw = clean_text(row.get(quantity_col, ""))
            unit_price_raw = clean_text(row.get(unit_price_col, ""))

            if not article_number and not description and not quantity_raw and not unit_price_raw:
                continue

            quantity = normalize_decimal_text(quantity_raw, "Quantity")
            unit_price_gross = normalize_decimal_text(unit_price_raw, "Unit Price")

            if not article_number and not description:
                continue
            if not article_number:
                raise ValueError("Missing Conrad Article-Nr. for one or more rows.")
            if not description:
                description = article_number

            rows.append(
                {
                    "product_name": build_product_name(article_number, description),
                    "description": description,
                    "quantity": quantity,
                    "unit_price": gross_to_net_price(unit_price_gross, vat_rate_percent),
                    "supplier_part_number": article_number,
                }
            )

    if not rows:
        raise ValueError("No item rows found in Conrad CSV.")
    return rows


def write_output_excel(output_path: Path, supplier_name: str, items: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["A2"] = supplier_name
    ws["A4"] = "Product name"
    ws["B4"] = "Description"
    ws["C4"] = "Quantity"
    ws["D4"] = "Unit price"
    ws["E4"] = "Supplier Part Number"

    row_idx = 5
    for item in items:
        ws.cell(row=row_idx, column=1, value=item["product_name"])
        ws.cell(row=row_idx, column=2, value=item["description"])
        ws.cell(row=row_idx, column=3, value=item["quantity"])
        ws.cell(row=row_idx, column=4, value=item["unit_price"])
        ws.cell(row=row_idx, column=5, value=item["supplier_part_number"])
        row_idx += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    scripts_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(
        description="Convert a Conrad cart CSV into Ariba import Excel format."
    )
    parser.add_argument("--csv", required=True, help="Path to Conrad cart CSV.")
    parser.add_argument("--out", required=True, help="Path to output .xlsx file.")
    parser.add_argument(
        "--supplier",
        default="Conrad",
        help="Supplier name written to cell A2 (default: Conrad).",
    )
    parser.add_argument(
        "--vat-xlsx",
        default=str(scripts_dir / "VATRate.xlsx"),
        help="Path to VATRate.xlsx (reads VAT rate from cell A2).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.csv).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()
    vat_xlsx_path = Path(args.vat_xlsx).expanduser().resolve()

    if not input_path.exists():
        print(f"Input CSV not found: {input_path}")
        return 2
    if not vat_xlsx_path.exists():
        print(f"VAT rate file not found: {vat_xlsx_path}")
        return 2

    try:
        vat_rate_percent = load_vat_rate_percent(vat_xlsx_path)
        items = parse_conrad_csv(input_path, vat_rate_percent)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"CSV parse error: {exc}")
        return 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(output_path, args.supplier, items)
    print(f"Wrote {len(items)} item(s) to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
