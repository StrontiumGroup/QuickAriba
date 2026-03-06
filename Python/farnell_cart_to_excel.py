#!/usr/bin/env python3
"""
Convert a Farnell shopping-cart CSV into the Excel format expected by ariba_excel_import.py.

Mapping:
- Farnell "Ordercode | Fabrikant / beschrijving" (max 80 chars) -> Excel "Product name"
- Farnell "Fabrikant / beschrijving"  -> Excel "Description"
- Farnell "Hoeveelheid"               -> Excel "Quantity"
- Farnell "Prijs per stuk"            -> Excel "Unit price"
- Farnell "Ordercode"                 -> Excel "Supplier Part Number"
"""

from __future__ import annotations

import argparse
import csv
import html
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook


def clean_text(value: str) -> str:
    return " ".join((value or "").strip().split())


def find_column(fieldnames: List[str], exact_name: str, prefix_name: str) -> str:
    if exact_name in fieldnames:
        return exact_name
    for name in fieldnames:
        if name.startswith(prefix_name):
            return name
    raise ValueError(f"CSV missing required column: '{exact_name}'")


def normalize_decimal_text(text: str, field_name: str) -> str:
    value = re.sub(r"[^0-9,.\-]", "", (text or "").strip()).replace(",", ".")
    if not value:
        raise ValueError(f"Missing numeric value for '{field_name}'.")
    try:
        decimal = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {field_name} value '{text}'.") from exc
    if decimal <= 0:
        raise ValueError(f"{field_name} must be > 0 (got '{text}').")
    return format(decimal.normalize(), "f")


def build_product_name(ordercode: str, description: str, max_len: int = 80) -> str:
    combined = f"{ordercode} | {description}"
    return combined[:max_len]


def parse_farnell_csv(csv_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames or []
        if not fieldnames:
            raise ValueError("CSV has no header row.")

        ordercode_col = find_column(fieldnames, "Ordercode", "Ordercode")
        description_col = find_column(
            fieldnames, "Fabrikant / beschrijving", "Fabrikant / beschrijving"
        )
        quantity_col = find_column(fieldnames, "Hoeveelheid", "Hoeveelheid")
        unit_price_col = find_column(fieldnames, "Prijs per stuk (&euro;)", "Prijs per stuk")

        for row in reader:
            ordercode = clean_text(row.get(ordercode_col, ""))
            if not ordercode:
                # Skip footer/subtotal lines.
                continue

            description_raw = html.unescape(row.get(description_col, "")).strip()
            description = clean_text(description_raw)
            quantity = normalize_decimal_text(row.get(quantity_col, ""), "Hoeveelheid")
            unit_price = normalize_decimal_text(
                row.get(unit_price_col, ""), "Prijs per stuk"
            )

            if not description:
                raise ValueError(f"Missing description for order code '{ordercode}'.")

            rows.append(
                {
                    "product_name": build_product_name(ordercode, description),
                    "description": description,
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "supplier_part_number": ordercode,
                }
            )

    if not rows:
        raise ValueError("No cart lines found in Farnell CSV.")
    return rows


def write_output_excel(
    output_path: Path, supplier_name: str, converted_rows: List[Dict[str, str]]
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["A2"] = supplier_name
    ws["A4"] = "Product name"
    ws["B4"] = "Description"
    ws["C4"] = "Quantity"
    ws["D4"] = "Unit price"
    ws["E4"] = "Supplier Part Number"

    excel_row = 5
    for row in converted_rows:
        ws.cell(row=excel_row, column=1, value=row["product_name"])
        ws.cell(row=excel_row, column=2, value=row["description"])
        ws.cell(row=excel_row, column=3, value=row["quantity"])
        ws.cell(row=excel_row, column=4, value=row["unit_price"])
        ws.cell(row=excel_row, column=5, value=row["supplier_part_number"])
        excel_row += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert a Farnell cart CSV into Ariba import Excel format."
    )
    parser.add_argument("--csv", required=True, help="Path to Farnell cart CSV.")
    parser.add_argument("--out", required=True, help="Path to output .xlsx file.")
    parser.add_argument(
        "--supplier",
        default="Farnell",
        help="Supplier name written to cell A2 (default: Farnell).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    csv_path = Path(args.csv).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()

    if not csv_path.exists():
        print(f"Input CSV not found: {csv_path}")
        return 2

    try:
        converted_rows = parse_farnell_csv(csv_path)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"CSV parse error: {exc}")
        return 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(output_path, args.supplier, converted_rows)
    print(f"Wrote {len(converted_rows)} item(s) to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
