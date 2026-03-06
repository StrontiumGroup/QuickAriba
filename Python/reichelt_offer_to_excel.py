#!/usr/bin/env python3
"""
Convert a Reichelt offer PDF into the Excel format expected by ariba_excel_import.py.

Mapping:
- Reichelt "Item No."    -> Excel "Product name"
- Reichelt "Description" -> Excel "Description"
- Quantity               -> Quantity
- Price                  -> Unit price

Ignored columns:
- Category of goods
- Price on all
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook


@dataclass
class OfferItem:
    item_no: str
    description: str
    quantity: str
    unit_price: str


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_euro_number(text: str) -> str:
    value = text.strip()
    value = re.sub(r"[^0-9,.\-]", "", value)
    value = value.replace(",", ".")
    return value


def load_vat_rate_percent(vat_xlsx_path: Path) -> Decimal:
    try:
        workbook = load_workbook(vat_xlsx_path, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Unable to read VAT rate file: {vat_xlsx_path}") from exc

    raw_value = workbook.active["A2"].value
    if raw_value is None:
        raise RuntimeError(f"VAT rate cell A2 is empty in: {vat_xlsx_path}")

    normalized = str(raw_value).strip().replace("%", "").replace(",", ".")
    try:
        vat_rate = Decimal(normalized)
    except InvalidOperation as exc:
        raise RuntimeError(
            f"VAT rate in A2 is not a valid number: {raw_value!r}"
        ) from exc

    if vat_rate < 0:
        raise RuntimeError(f"VAT rate in A2 must be >= 0, got: {vat_rate}")
    return vat_rate


def gross_to_net_price(gross_price: str, vat_rate_percent: Decimal) -> str:
    gross = Decimal(gross_price)
    divisor = Decimal("1") + (vat_rate_percent / Decimal("100"))
    net = (gross / divisor).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    net_text = format(net, "f").rstrip("0").rstrip(".")
    return net_text or "0"


def extract_text_with_pypdf(pdf_path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return ""

    reader = PdfReader(str(pdf_path))
    chunks: List[str] = []
    for page in reader.pages:
        # pypdf supports layout extraction in recent versions; fallback to default.
        try:
            text = page.extract_text(extraction_mode="layout")
        except TypeError:
            text = page.extract_text()
        chunks.append(text or "")
    return "\n".join(chunks)


def extract_text_with_pdftotext(pdf_path: Path) -> str:
    tool = shutil.which("pdftotext")
    if not tool:
        return ""
    try:
        result = subprocess.run(
            [tool, "-layout", "-enc", "UTF-8", str(pdf_path), "-"],
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout
    except Exception:
        return ""


def iter_table_lines(full_text: str) -> Iterable[str]:
    in_table = False
    for raw in full_text.splitlines():
        line = raw.rstrip()
        normalized = line.lower()

        if "item no." in normalized and "description" in normalized and "quantity" in normalized:
            in_table = True
            continue

        if not in_table:
            continue

        if "order value" in normalized:
            break

        if not line.strip():
            continue

        # Skip wrapped header fragments and non-row table text.
        if normalized.strip() in {"of goods", "incl. vat"}:
            continue
        if "of goods" in normalized and "incl. vat" in normalized:
            continue

        yield line


def split_item_and_description(lead: str) -> tuple[str, str] | tuple[None, None]:
    # Some PDFs keep a wide gap between item and description; others collapse them.
    lead_parts = re.split(r"\s{2,}", lead.strip())
    if len(lead_parts) >= 2:
        return clean_spaces(lead_parts[0]), clean_spaces(" ".join(lead_parts[1:]))

    # Fix missing space in extraction, e.g. "HLDeveloper" -> "HL Developer".
    normalized = re.sub(r"(?<=[A-Z0-9])(?=[A-Z][a-z])", " ", clean_spaces(lead))
    tokens = normalized.split()
    if len(tokens) < 2:
        return None, None

    # Item numbers are typically uppercase/digit tokens; description starts with lowercase text.
    split_at = None
    for idx, token in enumerate(tokens):
        if re.search(r"[a-z]", token):
            split_at = idx
            break
    if split_at is None or split_at == 0:
        return None, None

    item_no = clean_spaces(" ".join(tokens[:split_at]))
    description = clean_spaces(" ".join(tokens[split_at:]))
    if not item_no or not description:
        return None, None
    return item_no, description


def parse_row_line(line: str) -> OfferItem | None:
    parts = re.split(r"\s{2,}", line.strip())
    if len(parts) < 5:
        return None

    category = clean_spaces(parts[-4])
    quantity = clean_spaces(parts[-3])
    price = clean_spaces(parts[-2])
    lead = "  ".join(parts[:-4])

    if not category.isdigit() or not quantity.isdigit():
        return None

    item_no, description = split_item_and_description(lead)
    if not item_no or not description:
        return None

    unit_price = parse_euro_number(price)
    if not re.fullmatch(r"\d+(?:\.\d+)?", unit_price):
        return None

    return OfferItem(
        item_no=item_no,
        description=description,
        quantity=quantity,
        unit_price=unit_price,
    )


def looks_like_item_number_suffix(text: str) -> bool:
    return bool(
        re.fullmatch(r"[A-Z0-9][A-Z0-9 ./\-]*", text)
        and len(text.split()) <= 3
        and any(ch.isdigit() for ch in text)
    )


def parse_table_lines(lines: Iterable[str]) -> List[OfferItem]:
    items: List[OfferItem] = []

    for line in lines:
        parsed = parse_row_line(line)
        if parsed is not None:
            items.append(parsed)
            continue

        continuation = clean_spaces(line)
        if not continuation or not items:
            continue
        if looks_like_item_number_suffix(continuation):
            items[-1].item_no = f"{items[-1].item_no} {continuation}"
        else:
            items[-1].description = clean_spaces(f"{items[-1].description} {continuation}")

    return items


def write_output_excel(output_path: Path, supplier_name: str, items: List[OfferItem]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["B1"] = supplier_name

    ws["A3"] = "Product name"
    ws["B3"] = "Description"
    ws["C3"] = "Quantity"
    ws["D3"] = "Unit price"

    row = 4
    for item in items:
        # Mapping:
        # Reichelt Item No.    -> Product name
        # Reichelt Description -> Description
        ws.cell(row=row, column=1, value=item.item_no)
        ws.cell(row=row, column=2, value=item.description)
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=item.unit_price)
        row += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    scripts_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Convert a Reichelt offer PDF into Ariba import Excel format."
    )
    parser.add_argument("--pdf", required=True, help="Path to Reichelt offer PDF.")
    parser.add_argument(
        "--out",
        required=True,
        help="Path to output .xlsx file (for ariba_excel_import.py).",
    )
    parser.add_argument(
        "--supplier",
        default="Reichelt",
        help="Supplier name written to cell B1 (default: Reichelt).",
    )
    parser.add_argument(
        "--vat-xlsx",
        default=str(scripts_dir / "VATRate.xlsx"),
        help="Path to VATRate.xlsx (reads VAT rate from cell A2).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pdf_path = Path(args.pdf).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    vat_xlsx_path = Path(args.vat_xlsx).expanduser().resolve()

    if not pdf_path.exists():
        print(f"Input PDF not found: {pdf_path}")
        return 2

    text = extract_text_with_pypdf(pdf_path)
    if not text.strip():
        text = extract_text_with_pdftotext(pdf_path)

    if not text.strip():
        print(
            "Could not extract text from PDF. Install pypdf or ensure pdftotext is available."
        )
        return 2

    items = parse_table_lines(iter_table_lines(text))
    if not items:
        print("No offer rows found in PDF table.")
        return 1

    vat_rate_percent = load_vat_rate_percent(vat_xlsx_path)
    for item in items:
        item.unit_price = gross_to_net_price(item.unit_price, vat_rate_percent)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(out_path, args.supplier, items)

    print(f"Wrote {len(items)} item(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
