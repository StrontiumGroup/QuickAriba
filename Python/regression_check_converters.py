#!/usr/bin/env python3
"""
Run converter regression checks on the example supplier files.

The script regenerates Ariba-format Excel files from ExampleOffersFromSupplier
and compares rows against baseline files in ExampleOrdersForAriba.
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

from openpyxl import load_workbook


ARIBA_HEADERS = ["Product name", "Description", "Quantity", "Unit price"]


@dataclass(frozen=True)
class RegressionCase:
    name: str
    script_name: str
    input_flag: str
    input_file: str
    expected_file: str


CASES: Sequence[RegressionCase] = (
    RegressionCase(
        name="DigiKey",
        script_name="digikey_cart_to_excel.py",
        input_flag="--xlsx",
        input_file="ExampleDigikeyCart.xlsx",
        expected_file="orders_from_digikey.xlsx",
    ),
    RegressionCase(
        name="Farnell",
        script_name="farnell_cart_to_excel.py",
        input_flag="--csv",
        input_file="ExampleFarnellCart.csv",
        expected_file="orders_from_farnellcart.xlsx",
    ),
    RegressionCase(
        name="Mouser",
        script_name="mouser_cart_to_excel.py",
        input_flag="--xls",
        input_file="ExampleMouserCart.xls",
        expected_file="orders_from_mouser.xlsx",
    ),
    RegressionCase(
        name="Reichelt",
        script_name="reichelt_offer_to_excel.py",
        input_flag="--pdf",
        input_file="ExampleReicheltOffer.pdf",
        expected_file="orders_from_reicheltoffer.xlsx",
    ),
    RegressionCase(
        name="Thorlabs",
        script_name="thorlabs_cart_to_excel.py",
        input_flag="--csv",
        input_file="ExampleThorlabsCart.csv",
        expected_file="orders_from_thorlabs.xlsx",
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run converter regression checks against example files."
    )
    parser.add_argument(
        "--keep-temp",
        action="store_true",
        help="Keep generated files for inspection.",
    )
    return parser.parse_args()


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return f"{value:g}"

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    lines = [line for line in lines if line]
    return "\n".join(lines)


def detect_header_row(ws) -> int:
    for row in (3, 4):
        headers = [normalize_cell(ws.cell(row=row, column=col).value) for col in range(1, 5)]
        if headers == ARIBA_HEADERS:
            return row
    raise ValueError("Could not find Ariba header row (expected row 3 or 4).")


def read_ariba_rows(xlsx_path: Path) -> List[Tuple[str, str, str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header_row = detect_header_row(ws)
    rows: List[Tuple[str, str, str, str]] = []
    row_idx = header_row + 1

    while True:
        values = tuple(normalize_cell(ws.cell(row=row_idx, column=col).value) for col in range(1, 5))
        if all(not cell for cell in values):
            break
        rows.append(values)  # type: ignore[arg-type]
        row_idx += 1

    return rows


def run_converter(
    case: RegressionCase, python_exe: str, scripts_dir: Path, offers_dir: Path, out_file: Path
) -> None:
    command = [
        python_exe,
        str(scripts_dir / case.script_name),
        case.input_flag,
        str(offers_dir / case.input_file),
        "--out",
        str(out_file),
    ]
    result = subprocess.run(command, check=False, capture_output=True, text=True)
    if result.returncode != 0:
        stderr = (result.stderr or "").strip()
        stdout = (result.stdout or "").strip()
        details = stderr if stderr else stdout
        raise RuntimeError(f"Converter failed for {case.name}: {details}")


def format_first_difference(
    actual_rows: Iterable[Tuple[str, str, str, str]],
    expected_rows: Iterable[Tuple[str, str, str, str]],
) -> str:
    actual_list = list(actual_rows)
    expected_list = list(expected_rows)
    for idx, (actual, expected) in enumerate(zip(actual_list, expected_list), start=1):
        if actual != expected:
            return f"row {idx}: expected {expected}, got {actual}"
    if len(actual_list) != len(expected_list):
        return f"row count differs: expected {len(expected_list)}, got {len(actual_list)}"
    return "unknown difference"


def main() -> int:
    args = parse_args()
    scripts_dir = Path(__file__).resolve().parent
    repo_root = scripts_dir.parent
    offers_dir = repo_root / "ExampleOffersFromSupplier"
    expected_dir = repo_root / "ExampleOrdersForAriba"
    python_exe = sys.executable

    missing = []
    for case in CASES:
        if not (offers_dir / case.input_file).exists():
            missing.append(str(offers_dir / case.input_file))
        if not (expected_dir / case.expected_file).exists():
            missing.append(str(expected_dir / case.expected_file))
        if not (scripts_dir / case.script_name).exists():
            missing.append(str(scripts_dir / case.script_name))
    if missing:
        print("Missing required files:")
        for path in missing:
            print(f"- {path}")
        return 2

    keep_temp = args.keep_temp
    with tempfile.TemporaryDirectory(prefix="quickariba_regression_") as temp_dir:
        temp_path = Path(temp_dir)
        failures: List[str] = []

        for case in CASES:
            generated = temp_path / f"{case.name.lower()}_generated.xlsx"
            expected = expected_dir / case.expected_file
            try:
                run_converter(case, python_exe, scripts_dir, offers_dir, generated)
                actual_rows = read_ariba_rows(generated)
                expected_rows = read_ariba_rows(expected)
            except Exception as exc:
                failures.append(f"{case.name}: {exc}")
                continue

            if actual_rows != expected_rows:
                diff = format_first_difference(actual_rows, expected_rows)
                failures.append(f"{case.name}: output differs ({diff})")
            else:
                print(f"PASS {case.name} ({len(actual_rows)} rows)")

        if keep_temp:
            saved_dir = scripts_dir / "_regression_last_run"
            saved_dir.mkdir(parents=True, exist_ok=True)
            for file in temp_path.glob("*.xlsx"):
                shutil.copy2(file, saved_dir / file.name)
            print(f"\nSaved generated files to: {saved_dir}")

        if failures:
            print("\nRegression failures:")
            for failure in failures:
                print(f"- {failure}")
            return 1

        print("\nAll converter regression checks passed.")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
