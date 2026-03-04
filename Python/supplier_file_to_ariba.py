#!/usr/bin/env python3
"""
Detect supplier cart/offer format, convert if needed, then submit to Ariba.

Supported inputs:
- Ariba-ready .xlsx (A1/B1 + row 3 headers)
- Reichelt offer PDF
- Thorlabs cart CSV
- Farnell cart CSV
- DigiKey cart XLSX
- Mouser cart XLS
"""

from __future__ import annotations

import argparse
import csv
import os
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import xlrd
from openpyxl import load_workbook


ARIBA_HEADERS = ["Product name", "Description", "Quantity", "Unit price"]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_ariba_xlsx(path: Path) -> bool:
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
    except Exception:
        return False

    a1 = clean_text(ws["A1"].value).lower()
    if a1 != "supplier name":
        return False

    # New layout: metadata in rows 1-2, headers row 4, data row 5.
    new_headers = [clean_text(ws.cell(row=4, column=c).value) for c in range(1, 5)]
    if new_headers == ARIBA_HEADERS:
        first_data = [clean_text(ws.cell(row=5, column=c).value) for c in range(1, 5)]
        return any(first_data)

    # Legacy layout: headers row 3, data row 4.
    old_headers = [clean_text(ws.cell(row=3, column=c).value) for c in range(1, 5)]
    if old_headers == ARIBA_HEADERS:
        first_data = [clean_text(ws.cell(row=4, column=c).value) for c in range(1, 5)]
        return any(first_data)

    return False


def detect_digikey_xlsx(path: Path) -> bool:
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
    except Exception:
        return False

    row2 = [clean_text(ws.cell(row=2, column=c).value) for c in range(1, ws.max_column + 1)]
    required = {"Part Number", "Quantity", "Description", "Unit Price"}
    return required.issubset(set(row2))


def detect_csv_headers(path: Path) -> Optional[List[str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            headers = next(reader, [])
            return [clean_text(h) for h in headers]
    except Exception:
        return None


def detect_csv_supplier(path: Path) -> Optional[str]:
    headers = detect_csv_headers(path)
    if not headers:
        return None

    header_set = set(headers)
    thorlabs_needed = {"Item Number", "Quantity", "Description", "Unit Price", "URL"}
    if thorlabs_needed.issubset(header_set):
        return "thorlabs"

    if "Ordercode" in header_set and "Hoeveelheid" in header_set:
        return "farnell"

    return None


def detect_reichelt_pdf(path: Path) -> bool:
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        for page in reader.pages[:2]:
            page_text = page.extract_text() or ""
            text += page_text.lower()
    except Exception:
        pass

    if not text:
        tool = shutil.which("pdftotext")
        if tool:
            try:
                result = subprocess.run(
                    [tool, "-layout", "-enc", "UTF-8", str(path), "-"],
                    check=False,
                    capture_output=True,
                    text=True,
                )
                text = (result.stdout or "").lower()
            except Exception:
                text = ""

    return "reichelt" in text and "item no." in text and "description" in text


def detect_mouser_xls(path: Path) -> bool:
    try:
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
    except Exception:
        return False

    required = {"Mouser-nr", "Omschrijving", "Besteld aantal", "Prijs (EUR)"}
    for r in range(min(sheet.nrows, 30)):
        row = {clean_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)}
        if required.issubset(row):
            return True
    return False


def detect_input_kind(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        if detect_reichelt_pdf(path):
            return "reichelt_pdf"
        return "unknown_pdf"
    if suffix in {".csv", ".cvs"}:
        supplier = detect_csv_supplier(path)
        if supplier == "thorlabs":
            return "thorlabs_csv"
        if supplier == "farnell":
            return "farnell_csv"
        return "unknown_csv"
    if suffix == ".xlsx":
        if detect_ariba_xlsx(path):
            return "ariba_xlsx"
        if detect_digikey_xlsx(path):
            return "digikey_xlsx"
        return "unknown_xlsx"
    if suffix == ".xls":
        if detect_mouser_xls(path):
            return "mouser_xls"
        return "unknown_xls"
    return "unknown"


def default_output_path(input_path: Path, scripts_dir: Path) -> Path:
    stem = input_path.stem
    if stem.lower().startswith("example"):
        stem = stem[7:] or input_path.stem
    filename = f"orders_from_{stem.lower()}.xlsx"

    parent = input_path.parent
    if parent.name == "ExampleOffersFromSupplier":
        candidate = parent.parent / "ExampleOrdersForAriba"
        if candidate.exists():
            return candidate / filename
    return scripts_dir / filename


def run_command(command: List[str]) -> None:
    print("Running:", " ".join(f'"{p}"' if " " in p else p for p in command))
    env = os.environ.copy()
    env.setdefault("NODE_NO_WARNINGS", "1")
    result = subprocess.run(command, check=False, env=env)
    if result.returncode != 0:
        raise RuntimeError(f"Command failed with exit code {result.returncode}.")


def build_conversion_command(
    kind: str, input_path: Path, output_path: Path, python_exe: str, base_dir: Path
) -> Optional[List[str]]:
    converters: Dict[str, Tuple[str, str]] = {
        "reichelt_pdf": ("reichelt_offer_to_excel.py", "--pdf"),
        "thorlabs_csv": ("thorlabs_cart_to_excel.py", "--csv"),
        "farnell_csv": ("farnell_cart_to_excel.py", "--csv"),
        "digikey_xlsx": ("digikey_cart_to_excel.py", "--xlsx"),
        "mouser_xls": ("mouser_cart_to_excel.py", "--xls"),
    }
    if kind not in converters:
        return None

    script_name, arg_name = converters[kind]
    script_path = base_dir / script_name
    return [
        python_exe,
        str(script_path),
        arg_name,
        str(input_path),
        "--out",
        str(output_path),
    ]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Detect supplier cart/offer file, convert if needed, and run Ariba import."
    )
    parser.add_argument(
        "--input",
        nargs="?",
        const="",
        help=(
            "Path to supplier file or Ariba-ready XLSX. "
            "If omitted, the script looks for 'order.*' in the project root folder."
        ),
    )
    parser.add_argument(
        "--converted-out",
        help="Optional output .xlsx path when conversion is needed.",
    )
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222")
    parser.add_argument("--page-contains", default="ariba")
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug output in the downstream Ariba importer.",
    )
    parser.add_argument(
        "--detect-only",
        action="store_true",
        help="Only detect type and planned actions; do not run conversion/import.",
    )
    return parser.parse_args()


def normalize_input_arg(raw_input: Optional[str]) -> Optional[str]:
    if raw_input is None:
        return None
    value = raw_input.strip()
    if not value:
        return None
    if value in {'""', "''"}:
        return None
    return value


def auto_detect_order_input(root_dir: Path) -> Path:
    candidates = [p for p in root_dir.glob("order.*") if p.is_file()]
    if not candidates:
        raise FileNotFoundError(
            f"No input was provided and no 'order.*' file was found in: {root_dir}"
        )
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(
            "No --input provided; multiple 'order.*' files found. "
            f"Using newest: {candidates[0].name}"
        )
    else:
        print(f"No --input provided; using: {candidates[0].name}")
    return candidates[0]


def main() -> int:
    args = parse_args()
    scripts_dir = Path(__file__).resolve().parent
    project_root = scripts_dir.parent
    normalized_input = normalize_input_arg(args.input)
    if normalized_input:
        input_path = Path(normalized_input).expanduser().resolve()
    else:
        try:
            input_path = auto_detect_order_input(project_root)
        except FileNotFoundError as exc:
            print(exc)
            return 2
    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 2
    kind = detect_input_kind(input_path)
    print(f"Detected input type: {kind}")

    if kind.startswith("unknown"):
        print("Unsupported or unknown input format.")
        print("Supported: Ariba .xlsx, Reichelt .pdf, Thorlabs .csv, Farnell .csv, DigiKey .xlsx, Mouser .xls")
        return 2

    python_exe = sys.executable
    converted_path = (
        Path(args.converted_out).expanduser().resolve()
        if args.converted_out
        else default_output_path(input_path, scripts_dir)
    )

    conversion_cmd = build_conversion_command(
        kind, input_path, converted_path, python_exe, scripts_dir
    )
    ariba_input_path = input_path

    if args.detect_only:
        if conversion_cmd:
            print(f"Will convert to: {converted_path}")
            print("Conversion command:", " ".join(conversion_cmd))
            print("Then run ariba import with converted file.")
        else:
            print("Input already in Ariba format; no conversion needed.")
        return 0

    if conversion_cmd:
        converted_path.parent.mkdir(parents=True, exist_ok=True)
        run_command(conversion_cmd)
        ariba_input_path = converted_path
    else:
        ariba_input_path = input_path

    ariba_cmd = [
        python_exe,
        str(scripts_dir / "ariba_excel_import.py"),
        "--excel",
        str(ariba_input_path),
        "--cdp-url",
        args.cdp_url,
        "--page-contains",
        args.page_contains,
    ]
    if args.debug:
        ariba_cmd.append("--debug")
    run_command(ariba_cmd)
    return 0


if __name__ == "__main__":
    sys.exit(main())
