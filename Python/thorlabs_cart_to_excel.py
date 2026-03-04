#!/usr/bin/env python3
"""
Convert a Thorlabs cart CSV into the Excel format expected by ariba_excel_import.py.

Requested mapping:
- Thorlabs "Item Number" -> Excel "Product name"
- Excel "Description"    -> "<Thorlabs Description>\\nURL: <Thorlabs URL>"
- Thorlabs "Quantity"    -> Excel "Quantity"
- Thorlabs "Unit Price"  -> Excel "Unit price"
"""

from __future__ import annotations

import argparse
import csv
import html
import re
import sys
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook


REQUIRED_COLUMNS = ["Item Number", "Quantity", "Description", "Unit Price", "URL"]


def clean_text(value: str) -> str:
    return " ".join((value or "").strip().split())


def normalize_price_eur(value: str) -> str:
    # Example: "25,06 €" -> "25.06"
    cleaned = re.sub(r"[^0-9,.\-]", "", value or "")
    cleaned = cleaned.replace(",", ".")
    return cleaned


def build_description(description: str, url: str) -> str:
    desc = html.unescape(description or "").strip()
    url = (url or "").strip()
    if url:
        return f"{desc}\nURL: {url}"
    return desc


def is_data_row(row: Dict[str, str]) -> bool:
    return bool(clean_text(row.get("Item Number", "")))


def parse_thorlabs_csv(csv_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        missing = [c for c in REQUIRED_COLUMNS if c not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(f"CSV is missing required columns: {', '.join(missing)}")

        for row in reader:
            if not is_data_row(row):
                continue

            item_number = clean_text(row["Item Number"])
            quantity = clean_text(row["Quantity"])
            unit_price = normalize_price_eur(row["Unit Price"])
            description = build_description(row.get("Description", ""), row.get("URL", ""))

            if not item_number or not quantity or not unit_price:
                continue

            rows.append(
                {
                    "product_name": item_number,
                    "description": description,
                    "quantity": quantity,
                    "unit_price": unit_price,
                }
            )

    if not rows:
        raise ValueError("No data rows found in CSV.")
    return rows


def write_output_excel(output_path: Path, supplier_name: str, rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["B1"] = supplier_name

    ws["A3"] = "Product name"
    ws["B3"] = "Description"
    ws["C3"] = "Quantity"
    ws["D3"] = "Unit price"

    r = 4
    for row in rows:
        ws.cell(row=r, column=1, value=row["product_name"])
        ws.cell(row=r, column=2, value=row["description"])
        ws.cell(row=r, column=3, value=row["quantity"])
        ws.cell(row=r, column=4, value=row["unit_price"])
        r += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert a Thorlabs cart CSV into Ariba import Excel format."
    )
    parser.add_argument("--csv", required=True, help="Path to Thorlabs cart CSV.")
    parser.add_argument("--out", required=True, help="Path to output .xlsx file.")
    parser.add_argument(
        "--supplier",
        default="Thorlabs",
        help="Supplier name written to cell B1 (default: Thorlabs).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    csv_path = Path(args.csv).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()

    if not csv_path.exists():
        print(f"Input CSV not found: {csv_path}")
        return 2

    try:
        rows = parse_thorlabs_csv(csv_path)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"CSV parse error: {exc}")
        return 2

    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(out_path, args.supplier, rows)
    print(f"Wrote {len(rows)} item(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
