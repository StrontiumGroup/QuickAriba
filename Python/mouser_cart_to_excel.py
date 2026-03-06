#!/usr/bin/env python3
"""
Convert a Mouser cart XLS file into the Excel format expected by ariba_excel_import.py.

Mapping:
- Mouser "Mouser-nr | Omschrijving" (max 80 chars) -> Excel "Product name"
- Mouser "Omschrijving"   -> Excel "Description"
- Mouser "Besteld aantal" -> Excel "Quantity"
- Mouser "Prijs (EUR)"    -> Excel "Unit price"
- Mouser "Mouser-nr"      -> Excel "Supplier Part Number"
"""

from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List

import xlrd
from openpyxl import Workbook


REQUIRED_HEADERS = ["Mouser-nr", "Omschrijving", "Besteld aantal", "Prijs (EUR)"]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_decimal_text(value: object, field_name: str) -> str:
    text = clean_text(value)
    text = re.sub(r"[^0-9,.\-]", "", text).replace(",", ".")
    if not text:
        raise ValueError(f"Missing value for '{field_name}'.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {field_name} value '{value}'.") from exc
    if number <= 0:
        raise ValueError(f"{field_name} must be > 0 (got '{value}').")
    return format(number.normalize(), "f")


def build_product_name(mouser_number: str, description: str, max_len: int = 80) -> str:
    combined = f"{mouser_number} | {description}"
    return combined[:max_len]


def detect_header_row(sheet: xlrd.sheet.Sheet) -> int:
    for r in range(min(sheet.nrows, 30)):
        row = [clean_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        if all(header in row for header in REQUIRED_HEADERS):
            return r
    raise ValueError("Could not find Mouser header row in XLS.")


def build_column_map(sheet: xlrd.sheet.Sheet, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for c in range(sheet.ncols):
        header = clean_text(sheet.cell_value(header_row, c))
        if header in REQUIRED_HEADERS:
            header_map[header] = c

    missing = [h for h in REQUIRED_HEADERS if h not in header_map]
    if missing:
        raise ValueError(f"Missing required Mouser columns: {', '.join(missing)}")
    return header_map


def parse_mouser_xls(path: Path) -> List[Dict[str, str]]:
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)

    header_row = detect_header_row(sheet)
    col = build_column_map(sheet, header_row)

    rows: List[Dict[str, str]] = []
    for r in range(header_row + 1, sheet.nrows):
        product_name = clean_text(sheet.cell_value(r, col["Mouser-nr"]))
        description = clean_text(sheet.cell_value(r, col["Omschrijving"]))
        quantity_raw = sheet.cell_value(r, col["Besteld aantal"])
        unit_price_raw = sheet.cell_value(r, col["Prijs (EUR)"])

        if (
            product_name == ""
            and description == ""
            and clean_text(quantity_raw) == ""
            and clean_text(unit_price_raw) == ""
        ):
            continue

        if not product_name:
            # Skip totals/footer rows or notes.
            continue

        quantity = normalize_decimal_text(quantity_raw, "Besteld aantal")
        unit_price = normalize_decimal_text(unit_price_raw, "Prijs (EUR)")
        if not description:
            description = product_name

        rows.append(
            {
                "product_name": build_product_name(product_name, description),
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
                "supplier_part_number": product_name,
            }
        )

    if not rows:
        raise ValueError("No item rows found in Mouser XLS.")
    return rows


def write_output_excel(output_path: Path, supplier_name: str, items: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["A2"] = supplier_name
    ws["A4"] = "Product name"
    ws["B4"] = "Description"
    ws["C4"] = "Quantity"
    ws["D4"] = "Unit price"
    ws["E4"] = "Supplier Part Number"

    row_idx = 5
    for item in items:
        ws.cell(row=row_idx, column=1, value=item["product_name"])
        ws.cell(row=row_idx, column=2, value=item["description"])
        ws.cell(row=row_idx, column=3, value=item["quantity"])
        ws.cell(row=row_idx, column=4, value=item["unit_price"])
        ws.cell(row=row_idx, column=5, value=item["supplier_part_number"])
        row_idx += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert Mouser cart XLS into Ariba import Excel format."
    )
    parser.add_argument("--xls", required=True, help="Path to Mouser cart .xls file.")
    parser.add_argument("--out", required=True, help="Path to output .xlsx file.")
    parser.add_argument(
        "--supplier",
        default="Mouser",
        help="Supplier name written to cell A2 (default: Mouser).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.xls).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()

    if not input_path.exists():
        print(f"Input XLS not found: {input_path}")
        return 2

    try:
        items = parse_mouser_xls(input_path)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"Workbook parse error: {exc}")
        return 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(output_path, args.supplier, items)
    print(f"Wrote {len(items)} item(s) to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
