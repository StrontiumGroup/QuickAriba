#!/usr/bin/env python3
"""
Import non-catalog line items from Excel into SAP Ariba Buying.

Expected order Excel layout:
- A1: Supplier name
- B1: Supplier value (for example "Reichelt" or "Thorlabs")
- Row 3 headers: Product name | Description | Quantity | Unit price
- Data starts at row 4

Additional required metadata file (searched next to this script first, then one folder up):
- PaymentAndShipping.xlsx
  - A1: Need-by-date, A2: value
  - B1: Deliver to,  B2: value
  - C1: WBS,         C2: value

The script uses Playwright and attaches to an already open Chromium browser
instance via CDP. This is intentional so users can complete SSO/2FA manually.
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import datetime
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Locator, Page, TimeoutError, sync_playwright


HEADER_ROW = 3
DATA_START_ROW = 4
REQUIRED_HEADERS = ["Product name", "Description", "Quantity", "Unit price"]
ITEM_SHIPTO_TARGET_PREFIX = "0006 (UvA"
DEBUG = False
PAYMENT_SHIPPING_DEFAULT_FILENAME = "PaymentAndShipping.xlsx"


@dataclass
class OrderMeta:
    supplier_name: str
    need_by_date: str
    deliver_to: str
    wbs: str


@dataclass
class ItemRow:
    row_number: int
    product_name: str
    description: str
    quantity: str
    unit_price: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def default_payment_shipping_path(script_dir: Path) -> Path:
    local_path = script_dir / PAYMENT_SHIPPING_DEFAULT_FILENAME
    if local_path.exists():
        return local_path
    return script_dir.parent / PAYMENT_SHIPPING_DEFAULT_FILENAME


def dbg(message: str) -> None:
    if DEBUG:
        print(f"[DEBUG] {message}")


def normalize_need_by_date(value: str) -> str:
    text = clean_text(value)
    if not text:
        return text

    # Already likely in accepted format like "March 20, 2026".
    if re.search(r"[A-Za-z]", text):
        return text

    patterns = ["%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"]
    for pattern in patterns:
        try:
            dt = datetime.strptime(text, pattern)
            return f"{dt.strftime('%B')} {dt.day}, {dt.year}"
        except ValueError:
            continue
    return text


def normalize_number_text(value: object, field_name: str, row_number: int) -> str:
    text = clean_text(value)
    if not text:
        raise ValueError(f"Row {row_number}: missing {field_name}.")
    normalized = text.replace(",", ".")
    try:
        decimal = Decimal(normalized)
    except InvalidOperation as exc:
        raise ValueError(
            f"Row {row_number}: invalid {field_name} '{text}'."
        ) from exc
    if decimal <= 0:
        raise ValueError(f"Row {row_number}: {field_name} must be > 0.")
    return format(decimal.normalize(), "f")


def load_items_from_excel(excel_path: Path) -> tuple[OrderMeta, List[ItemRow]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    supplier_name = ""
    need_by_date = ""
    deliver_to = ""
    wbs = ""
    header_row = HEADER_ROW
    data_start_row = DATA_START_ROW

    # Order file uses the original layout only.
    a1 = clean_text(ws["A1"].value).lower()
    if a1 != "supplier name":
        raise ValueError("Cell A1 must contain 'Supplier name'.")
    supplier_name = clean_text(ws["B1"].value)

    if not supplier_name:
        raise ValueError("Supplier value is empty.")

    headers = [clean_text(ws.cell(row=header_row, column=i).value) for i in range(1, 5)]
    if headers != REQUIRED_HEADERS:
        raise ValueError(
            f"Row {header_row} headers must be exactly: {', '.join(REQUIRED_HEADERS)}."
        )

    items: List[ItemRow] = []
    current_row = data_start_row
    while True:
        values = [ws.cell(row=current_row, column=i).value for i in range(1, 5)]
        if all(clean_text(v) == "" for v in values):
            break

        product_name = clean_text(values[0])
        description = clean_text(values[1])
        if not product_name:
            raise ValueError(f"Row {current_row}: Product name is empty.")
        if not description:
            raise ValueError(f"Row {current_row}: Description is empty.")

        quantity = normalize_number_text(values[2], "Quantity", current_row)
        unit_price = normalize_number_text(values[3], "Unit price", current_row)

        items.append(
            ItemRow(
                row_number=current_row,
                product_name=product_name,
                description=description,
                quantity=quantity,
                unit_price=unit_price,
            )
        )
        current_row += 1

    if not items:
        raise ValueError(f"No item rows found starting at row {data_start_row}.")

    return OrderMeta(supplier_name=supplier_name, need_by_date=need_by_date, deliver_to=deliver_to, wbs=wbs), items


def load_payment_and_shipping_meta(meta_path: Path, supplier_name: str) -> OrderMeta:
    wb = load_workbook(meta_path, data_only=True)
    ws = wb.active

    a1 = clean_text(ws["A1"].value).lower()
    b1 = clean_text(ws["B1"].value).lower()
    c1 = clean_text(ws["C1"].value).lower()

    if a1 != "need-by-date":
        raise ValueError("PaymentAndShipping.xlsx: cell A1 must be 'Need-by-date'.")
    if b1 != "deliver to":
        raise ValueError("PaymentAndShipping.xlsx: cell B1 must be 'Deliver to'.")
    if c1 != "wbs":
        raise ValueError("PaymentAndShipping.xlsx: cell C1 must be 'WBS'.")

    need_by_date = clean_text(ws["A2"].value)
    deliver_to = clean_text(ws["B2"].value)
    wbs = clean_text(ws["C2"].value)

    if not need_by_date:
        raise ValueError("PaymentAndShipping.xlsx: cell A2 (Need-by-date) is empty.")
    if not deliver_to:
        raise ValueError("PaymentAndShipping.xlsx: cell B2 (Deliver to) is empty.")
    if not wbs:
        raise ValueError("PaymentAndShipping.xlsx: cell C2 (WBS) is empty.")

    return OrderMeta(
        supplier_name=supplier_name,
        need_by_date=need_by_date,
        deliver_to=deliver_to,
        wbs=wbs,
    )


def wait_for_non_catalog_page(page: Page) -> None:
    page.get_by_text("Non-catalog request", exact=False).wait_for(timeout=20000)


def select_fixed_category(page: Page) -> None:
    category_target = "900006 (Laboratory \u2013 Items and disposables)"
    already_selected = page.get_by_text("900006 (Laboratory", exact=False).first
    if already_selected.count() > 0:
        try:
            if already_selected.is_visible():
                return
        except PlaywrightError:
            pass

    category_field = page.get_by_text("Choose a category", exact=False).first
    for _ in range(3):
        category_field.wait_for(timeout=2500)
        category_field.click()

        # Ariba sometimes lazily renders category rows only after pointer movement.
        try:
            clear_selection = page.get_by_text("Clear selection", exact=False).first
            if clear_selection.count() > 0:
                clear_selection.hover(timeout=500)
        except PlaywrightError:
            pass

        option = page.get_by_text(category_target, exact=True).first
        try:
            option.wait_for(timeout=1200)
            option.click(timeout=1200)
            return
        except PlaywrightError:
            continue

    # Final fallback: verify if selected despite missed click timing.
    page.get_by_text("900006 (Laboratory", exact=False).first.wait_for(timeout=5000)


def fill_product_name(page: Page, value: str) -> None:
    input_box = page.locator("xpath=//*[contains(normalize-space(.), 'Product name')]/following::input[1]").first
    input_box.wait_for(timeout=5000)
    input_box.fill(value)


def fill_description(page: Page, value: str) -> None:
    textareas = page.locator("textarea")
    if textareas.count() == 0:
        raise RuntimeError("Unable to locate Description textarea.")
    textareas.nth(0).fill(value)


def first_visible_or_none(locators: List) -> Optional:
    for locator in locators:
        try:
            count = locator.count()
        except PlaywrightError:
            continue
        for i in range(count):
            candidate = locator.nth(i)
            try:
                if candidate.is_visible():
                    return candidate
            except PlaywrightError:
                continue
    return None


def first_visible_text(page: Page, text: str):
    matches = page.get_by_text(text, exact=False)
    try:
        count = matches.count()
    except PlaywrightError:
        return None
    for i in range(count):
        candidate = matches.nth(i)
        try:
            if candidate.is_visible():
                return candidate
        except PlaywrightError:
            continue
    return None


def wait_for_visible_text(page: Page, text: str, timeout_ms: int = 10000) -> None:
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        if first_visible_text(page, text) is not None:
            return
        page.wait_for_timeout(200)
    raise TimeoutError(f"Timed out waiting for visible text: {text}")


def supplier_name_token(supplier_name: str) -> str:
    parts = [p for p in supplier_name.split() if p]
    return parts[0] if parts else supplier_name


def fill_quantity(page: Page, value: str) -> None:
    quantity_input = first_visible_or_none(
        [
            page.locator(
                "input:visible[aria-label*='quantity' i]:not(.shopping-cart-line-item-quantity-input)"
            ),
            page.locator(
                "xpath=//button[normalize-space()='-']/following::input[1]"
            ),
            page.locator(
                "xpath=//*[contains(normalize-space(.), 'Quantity')]/following::input[@name!='searchInGB'][1]"
            ),
        ]
    )
    if quantity_input is None:
        raise RuntimeError("Unable to locate visible Quantity field.")
    quantity_input.fill(value)
    quantity_input.evaluate(
        "el => el.dispatchEvent(new Event('change', { bubbles: true }))"
    )


def fill_unit_price(page: Page, value: str) -> None:
    price_input = first_visible_or_none(
        [
            page.locator("input:visible[aria-label*='price' i]"),
            page.locator(
                "xpath=//*[contains(normalize-space(.), 'Unit price')]/following::input[@name!='searchInGB'][1]"
            ),
            page.locator("xpath=//*[normalize-space(text())='EUR']/preceding::input[1]"),
        ]
    )
    if price_input is None:
        raise RuntimeError("Unable to locate visible Unit price field.")
    price_input.fill(value)
    price_input.evaluate(
        "el => el.dispatchEvent(new Event('change', { bubbles: true }))"
    )


def open_supplier_dialog(page: Page) -> None:
    page.get_by_text("View all suppliers", exact=False).first.click()
    page.get_by_text("Select a supplier", exact=True).wait_for(timeout=10000)
    page.locator("input[placeholder='Search']:visible").first.wait_for(timeout=10000)


def select_supplier(page: Page, supplier_name: str) -> None:
    if is_supplier_already_selected(page, supplier_name):
        return

    open_supplier_dialog(page)

    dialog = page.locator("xpath=//*[contains(normalize-space(.), 'Select a supplier')]/ancestor::*[1]").first
    search_input = dialog.locator("input[placeholder='Search']:visible").first
    search_input.click()
    search_input.fill("")
    search_input.fill(supplier_name)
    search_input.press("Enter")
    # Some tenants only apply the filter when search icon is clicked.
    search_icon = dialog.locator(
        "xpath=.//*[contains(@class,'search') or contains(@aria-label,'Search') or normalize-space(text())='🔍']"
    ).first
    if search_icon.count() > 0:
        try:
            search_icon.click(timeout=700)
        except PlaywrightError:
            pass

    supplier_tokens = [t for t in re.split(r"\s+", supplier_name) if t]
    token = supplier_tokens[0] if supplier_tokens else supplier_name

    row_candidates = dialog.locator("tr, [role='row'], .fd-table__row, td, [role='cell'], .fd-table__cell, .fdp-table__cell")
    try:
        row_candidates.first.wait_for(timeout=3000)
    except TimeoutError:
        pass

    def first_visible(locator):
        try:
            count = locator.count()
        except PlaywrightError:
            return None
        for idx in range(count):
            cand = locator.nth(idx)
            try:
                if cand.is_visible():
                    return cand
            except PlaywrightError:
                continue
        return None

    target = None
    deadline = time.monotonic() + 12.0
    while time.monotonic() < deadline and target is None:
        # First try row-level matching.
        target = first_visible(
            dialog.locator(
                "tr, [role='row'], .fd-table__row",
                has_text=re.compile(rf"\b{re.escape(supplier_name)}\b", re.IGNORECASE),
            )
        )
        if target is None and token:
            target = first_visible(
                dialog.locator(
                    "tr, [role='row'], .fd-table__row",
                    has_text=re.compile(re.escape(token), re.IGNORECASE),
                )
            )

        # Fallback: table-cell level matching for newer FD table implementations.
        if target is None:
            target = first_visible(
                dialog.locator(
                    "td, [role='cell'], .fd-table__cell, .fdp-table__cell",
                    has_text=re.compile(rf"\b{re.escape(supplier_name)}\b", re.IGNORECASE),
                )
            )
        if target is None and token:
            target = first_visible(
                dialog.locator(
                    "td, [role='cell'], .fd-table__cell, .fdp-table__cell",
                    has_text=re.compile(re.escape(token), re.IGNORECASE),
                )
            )
        if target is not None:
            break

        # Last fallback: any visible text node in dialog containing supplier token.
        if token:
            token_hit = first_visible(
                dialog.locator(
                    "xpath=.//*[contains(translate(normalize-space(.), "
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), "
                    f"'{token.lower()}')]"
                )
            )
            if token_hit is not None:
                target = token_hit
                break

        # Retry trigger: click search icon/refresh icon if present.
        try:
            search_icon.click(timeout=600)
        except PlaywrightError:
            pass
        refresh_icon = dialog.locator(
            "xpath=.//*[contains(@class,'refresh') or contains(@aria-label,'Refresh') or contains(normalize-space(.), '↻')]"
        ).first
        if refresh_icon.count() > 0:
            try:
                refresh_icon.click(timeout=600)
            except PlaywrightError:
                pass
        page.wait_for_timeout(500)

    if target is None:
        raise RuntimeError(f"Supplier search found no match for '{supplier_name}'.")

    # Click row ancestor when possible, otherwise click the cell itself.
    try:
        row = target.locator("xpath=ancestor::*[self::tr or @role='row' or contains(@class,'fd-table__row')][1]").first
        if row.count() > 0 and row.is_visible():
            row.click()
        else:
            target.click()
    except PlaywrightError:
        target.click()
    select_button = dialog.get_by_role("button", name="Select").locator(":visible").first
    select_button.click()
    # Accept either explicit selected badge or visible supplier name in chosen-supplier card.
    token = supplier_name_token(supplier_name)
    deadline = time.monotonic() + 10.0
    while time.monotonic() < deadline:
        if is_supplier_already_selected(page, supplier_name):
            return
        if token and first_visible_text(page, token) is not None:
            return
        page.wait_for_timeout(200)
    raise TimeoutError("Timed out confirming supplier selection.")


def verify_supplier_selected(page: Page, supplier_name: str) -> None:
    supplier_section = page.locator("xpath=//*[contains(normalize-space(.), 'Chosen supplier')]/ancestor::*[1]").first
    supplier_section.wait_for(timeout=8000)
    if first_visible_text(page, "Selected") is not None:
        return
    token = supplier_name_token(supplier_name)
    if token and first_visible_text(page, token) is not None:
        return
    raise RuntimeError("Supplier card is present but could not confirm selected supplier.")


def is_supplier_already_selected(page: Page, supplier_name: str) -> bool:
    if first_visible_text(page, "Select a supplier") is not None:
        return False
    supplier_panel = first_visible_text(page, "Chosen supplier")
    if supplier_panel is None:
        return False
    if first_visible_text(page, "Selected") is not None:
        return True
    token = supplier_name_token(supplier_name)
    return bool(token and first_visible_text(page, token) is not None)


def click_add_to_cart(page: Page) -> None:
    page.get_by_role("button", name="Add to cart").first.click()
    # Fast wait for immediate post-click state change.
    overlay = page.locator(
        "xpath=//*[contains(normalize-space(.), 'items in your cart')]"
    ).first
    try:
        overlay.wait_for(timeout=1200)
    except TimeoutError:
        try:
            page.get_by_role("button", name="Done").first.wait_for(timeout=1200)
        except TimeoutError:
            # Some UI states keep Add to cart visible; continue quickly.
            pass


def close_cart_overlay_if_present(page: Page) -> None:
    overlay_header = page.locator(
        "xpath=//*[contains(normalize-space(.), 'items in your cart')]"
    ).first
    if overlay_header.count() == 0:
        return

    close_candidates = [
        page.get_by_role(
            "button", name=re.compile(r"^(x|×|close)$", re.IGNORECASE)
        ).first,
        page.locator("button:has-text('X'):visible").first,
        page.locator(
            "xpath=//*[contains(normalize-space(.), 'items in your cart')]/ancestor::*[1]//button[1]"
        ).first,
    ]
    for close_btn in close_candidates:
        if close_btn.count() == 0:
            continue
        try:
            close_btn.click(timeout=1200)
            overlay_header.wait_for(state="hidden", timeout=1200)
            return
        except PlaywrightError:
            continue

    # Fallbacks if the close icon is not clickable.
    try:
        page.keyboard.press("Escape")
        overlay_header.wait_for(state="hidden", timeout=800)
        return
    except PlaywrightError:
        pass
    try:
        page.mouse.click(80, 220)
        overlay_header.wait_for(state="hidden", timeout=800)
    except PlaywrightError:
        pass


def create_new_item(page: Page) -> None:
    close_cart_overlay_if_present(page)

    menu_candidates = [
        page.locator("xpath=//button[normalize-space()='Done']/following::button[1]").first,
        page.get_by_role("button", name="...").first,
        page.locator("button:has-text('...')").first,
        page.locator("xpath=//*[normalize-space(text())='...']").first,
    ]

    clicked_menu = False
    for candidate in menu_candidates:
        try:
            candidate.click(timeout=450)
            clicked_menu = True
            break
        except PlaywrightError:
            continue

    if not clicked_menu:
        # Last-resort click: the menu is just to the right of "Done".
        try:
            done_btn = page.get_by_role("button", name="Done").first
            box = done_btn.bounding_box()
            if box is not None:
                page.mouse.click(box["x"] + box["width"] + 40, box["y"] + box["height"] / 2)
                clicked_menu = True
        except PlaywrightError:
            pass

    if not clicked_menu:
        raise RuntimeError("Unable to open the three-dot menu.")

    page.get_by_text("Create new", exact=True).first.wait_for(timeout=1200)
    page.get_by_text("Create new", exact=True).first.click()
    page.get_by_role("button", name="Add to cart").first.wait_for(timeout=2000)


def ensure_cart_overlay_open(page: Page) -> None:
    overlay_header = page.locator("xpath=//*[contains(normalize-space(.), 'items in your cart')]").first
    try:
        overlay_header.wait_for(timeout=1200)
        return
    except TimeoutError:
        pass

    # Fallback: click cart icon in header if overlay is not currently visible.
    cart_candidates = [
        page.locator("xpath=//*[contains(@aria-label,'cart') or contains(@title,'cart')]").first,
        page.locator("xpath=//*[contains(@class,'cart') and (self::button or self::a or @role='button')]").first,
    ]
    for candidate in cart_candidates:
        if candidate.count() == 0:
            continue
        try:
            candidate.click(timeout=1000)
            overlay_header.wait_for(timeout=2000)
            return
        except PlaywrightError:
            continue

    raise RuntimeError("Unable to open cart overlay.")


def click_checkout(page: Page) -> None:
    ensure_cart_overlay_open(page)
    checkout_candidates = [
        page.get_by_role("button", name="Check out").first,
        page.locator("button:has-text('Check out'):visible").first,
        page.locator("xpath=//*[self::button or self::a][contains(normalize-space(.), 'Check out')]").first,
        page.get_by_text("Check out", exact=False).first,
    ]

    clicked = False
    for candidate in checkout_candidates:
        if candidate.count() == 0:
            continue
        try:
            candidate.scroll_into_view_if_needed(timeout=800)
        except PlaywrightError:
            pass
        try:
            candidate.click(timeout=1200)
            clicked = True
            break
        except PlaywrightError:
            continue

    if not clicked:
        raise RuntimeError("Unable to click 'Check out' in cart overlay.")

    # Requisition page has this field at the top.
    page.get_by_text("Requisition title", exact=False).first.wait_for(timeout=20000)


def fill_input_if_needed(inp, target_value: str, page: Page) -> bool:
    try:
        current = clean_text(inp.input_value())
    except PlaywrightError:
        current = clean_text(inp.text_content())

    if current == clean_text(target_value):
        dbg(f"Field already has target value '{target_value}'.")
        return False

    dbg(f"Filling field from '{current}' to '{target_value}'.")
    inp.fill(target_value)
    inp.evaluate("el => el.dispatchEvent(new Event('change', { bubbles: true }))")
    try:
        inp.press("Tab")
    except PlaywrightError:
        pass
    try:
        page.keyboard.press("Escape")
    except PlaywrightError:
        pass
    return True


def safe_text(locator: Locator, timeout_ms: int = 700) -> str:
    try:
        return clean_text(locator.text_content(timeout=timeout_ms))
    except PlaywrightError:
        return ""


def safe_attr(locator: Locator, name: str, timeout_ms: int = 500) -> str:
    try:
        value = locator.get_attribute(name, timeout=timeout_ms)
        return clean_text(value)
    except Exception:
        return ""


def fill_first_input_near_label(
    page: Page,
    label: str,
    value: str,
    min_y: float = -1,
    max_y: float = 10_000_000,
    require_change: bool = False,
) -> bool:
    # Prefer explicit control ids when available in this Ariba tenant.
    explicit_candidates = []
    if label == "Deliver To":
        explicit_candidates.append(
            page.locator("input[name='DeliverTo'][data-help-id*='-DeliverTo']:visible")
        )
    if label == "Need-by Date":
        explicit_candidates.append(
            page.locator("input[data-help-id*='NeedBy']:visible, input[data-help-id*='Need-by']:visible")
        )

    for locator in explicit_candidates:
        try:
            count = locator.count()
        except PlaywrightError:
            continue
        for i in range(count):
            inp = locator.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None:
                    continue
                if box["y"] <= min_y or box["y"] >= max_y:
                    continue
                changed = fill_input_if_needed(inp, value, page)
                return changed if require_change else True
            except PlaywrightError:
                continue

    labels = page.locator(f"xpath=//*[contains(normalize-space(.), '{label}')]")
    label_count = labels.count()
    for li in range(label_count):
        label_el = labels.nth(li)
        try:
            if not label_el.is_visible():
                continue
            lbox = label_el.bounding_box()
            if lbox is None:
                continue
            if lbox["y"] <= min_y or lbox["y"] >= max_y:
                continue
        except PlaywrightError:
            continue

        # Search a handful of nearby following inputs and pick one in the same visual row/area.
        nearby = label_el.locator("xpath=following::input[position()<=8]")
        nearby_count = nearby.count()
        for i in range(nearby_count):
            inp = nearby.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None:
                    continue
                if box["y"] <= min_y or box["y"] >= max_y:
                    continue
                # Keep input geometrically close to the label to avoid cross-filling neighboring fields.
                if abs(box["y"] - lbox["y"]) > 160:
                    continue
                if box["x"] + 30 < lbox["x"]:
                    continue
                # Keep left/right required fields from cross-filling each other.
                if label == "Need-by Date" and box["x"] > lbox["x"] + 260:
                    continue
                if label == "Deliver To" and box["x"] < lbox["x"] - 20:
                    continue
                changed = fill_input_if_needed(inp, value, page)
                return changed if require_change else True
            except PlaywrightError:
                continue
    return False


def fill_requisition_header_fields(page: Page, meta: OrderMeta) -> None:
    items_header = first_visible_text(page, "Items (")
    items_y = -1.0
    if items_header is not None:
        try:
            box = items_header.bounding_box()
            if box:
                items_y = box["y"]
        except PlaywrightError:
            pass

    normalized_date = normalize_need_by_date(meta.need_by_date)

    if normalized_date:
        ok = fill_requisition_header_field(
            page,
            label="Need-by Date",
            value=normalized_date,
            max_y=items_y if items_y > 0 else 10_000_000,
        )
        if not ok:
            raise RuntimeError("Could not fill requisition Need-by Date.")
        dbg("Filled requisition header Need-by Date.")

    if meta.deliver_to:
        ok = fill_requisition_header_field(
            page,
            label="Deliver To",
            value=meta.deliver_to,
            max_y=items_y if items_y > 0 else 10_000_000,
        )
        if not ok:
            raise RuntimeError("Could not fill requisition Deliver To.")
        # Ariba occasionally keeps the field empty after a seemingly successful fill.
        # If the inline validation still says "Deliver To must be set.", retry once or twice.
        for _ in range(2):
            page.wait_for_timeout(300)
            if not has_visible_deliver_to_required_error(page, max_y=items_y if items_y > 0 else 10_000_000):
                break
            dbg("Header Deliver To validation still visible; retrying fill.")
            ok = fill_requisition_header_field(
                page,
                label="Deliver To",
                value=meta.deliver_to,
                max_y=items_y if items_y > 0 else 10_000_000,
            )
            if not ok:
                break
        dbg("Filled requisition header Deliver To.")


def has_visible_deliver_to_required_error(page: Page, max_y: float) -> bool:
    msg = page.locator("text=/Deliver To\\s+must be set\\./i")
    try:
        count = msg.count()
    except PlaywrightError:
        return False
    for i in range(count):
        err = msg.nth(i)
        try:
            if not err.is_visible():
                continue
            box = err.bounding_box()
            if box is None:
                continue
            if box["y"] < max_y:
                return True
        except PlaywrightError:
            continue
    return False


def fill_requisition_header_field(page: Page, label: str, value: str, max_y: float) -> bool:
    # Strong, explicit selectors first (top requisition section only).
    explicit_candidates = []
    if label == "Deliver To":
        explicit_candidates.append(page.locator("input[name='DeliverTo'][data-help-id*='-DeliverTo']:visible"))
    if label == "Need-by Date":
        explicit_candidates.append(page.locator("input[data-help-id*='NeedBy']:visible, input[data-help-id*='Need-by']:visible"))

    for locator in explicit_candidates:
        try:
            count = locator.count()
        except PlaywrightError:
            continue
        for i in range(count):
            inp = locator.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None or box["y"] >= max_y:
                    continue
                fill_input_if_needed(inp, value, page)
                return True
            except PlaywrightError:
                continue

    # Restrict to top requisition section (above Items list) and ignore error/help text lines.
    label_candidates = page.get_by_text(label, exact=False)
    count = label_candidates.count()
    for i in range(count):
        lbl = label_candidates.nth(i)
        try:
            if not lbl.is_visible():
                continue
            lbox = lbl.bounding_box()
            if lbox is None:
                continue
            if lbox["y"] >= max_y:
                continue
            ltext = clean_text(lbl.text_content())
            ltext_low = ltext.lower()
            label_low = label.lower()
            if not ltext_low.startswith(label_low):
                continue
            if "must be set" in ltext_low or "empty" in ltext_low or "mypupcode" in ltext_low:
                continue
            if "deliverto" in ltext_low:
                continue
        except PlaywrightError:
            continue

        nearby = lbl.locator("xpath=following::input[position()<=6]")
        ncount = nearby.count()
        for j in range(ncount):
            inp = nearby.nth(j)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None:
                    continue
                if box["y"] >= max_y:
                    continue
                # Keep close to the label row to avoid cross-field writes.
                if abs(box["y"] - lbox["y"]) > 120:
                    continue
                if box["x"] + 20 < lbox["x"]:
                    continue
                # Extra guard to avoid swapping left/right fields.
                if label == "Need-by Date" and box["x"] > lbox["x"] + 260:
                    continue
                if label == "Deliver To" and box["x"] < lbox["x"] - 20:
                    continue
                dbg(f"Header field '{label}' using label text '{ltext}' at y={lbox['y']:.0f}.")
                fill_input_if_needed(inp, value, page)
                return True
            except PlaywrightError:
                continue
    return False


def click_visible_item_expanders(page: Page, min_y: float) -> int:
    clicked = 0
    expanders = page.locator(
        "xpath=("
        "//*[self::button or self::a or self::span or self::i]"
        "[normalize-space(text())='>' or normalize-space(text())='▸' or normalize-space(text())='▶' "
        "or contains(@aria-label,'Expand') or contains(@title,'Expand') "
        "or contains(@class,'arrow-right') or contains(@class,'chevron-right') or contains(@class,'icon-slim-arrow-right')]"
        "| //*[@role='button' and (contains(@aria-label,'Expand') or contains(@title,'Expand'))]"
        ")"
    )
    try:
        count = expanders.count()
    except PlaywrightError:
        count = 0
    dbg(f"Found expander candidates: {count}")
    for i in range(count):
        exp = expanders.nth(i)
        try:
            if not exp.is_visible():
                continue
            aria = clean_text(exp.get_attribute("aria-label")).lower()
            if "accounting" in aria:
                continue
            txt = clean_text(exp.text_content()).lower()
            if "accounting" in txt:
                continue
            in_accounting = exp.locator(
                "xpath=ancestor::*[contains(@class,'accounting-section') or contains(@aria-label,'Accounting')]"
            ).count()
            if in_accounting > 0:
                continue
            box = exp.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            # Item expanders are on the far-left side of each item row.
            if box["x"] > 180:
                continue
            exp.click(timeout=300)
            clicked += 1
        except PlaywrightError:
            continue
    dbg(f"Clicked expanders this pass: {clicked}")
    return clicked


def fill_item_fields_on_requisition(page: Page, meta: OrderMeta, expected_items: int) -> None:
    items_y = get_items_section_min_y(page)
    # Start from top of items so first rows are not missed.
    try:
        page.mouse.wheel(0, -5000)
    except PlaywrightError:
        pass

    # Need-by Date and Deliver To are filled once in header and then inherited by items.
    attempts_without_progress = 0
    zero_banner_rounds = 0
    filled_shipto = 0
    filled_wbs = 0
    if not meta.wbs:
        dbg("WBS is empty; skipping Accounting/WBS automation.")
    max_rounds = max(55, expected_items * 10)

    for round_idx in range(max_rounds):
        progress = 0
        round_start = time.monotonic()
        expanded = 0
        extra_shipto = 0
        extra_wbs = 0
        scroll_delta = 0
        dbg(
            f"Requisition item round start: no-progress={attempts_without_progress}, "
            f"filled_shipto={filled_shipto}, filled_wbs={filled_wbs}"
        )

        remaining_before = count_visible_item_error_banners(page, min_y=items_y)
        dbg(f"Visible item error banners: {remaining_before}")

        progress_shipto, progress_wbs = process_visible_items_with_errors(
            page,
            meta,
            min_y=items_y,
            include_all_visible_items=False,
        )
        filled_shipto += progress_shipto
        filled_wbs += progress_wbs
        progress += progress_shipto + progress_wbs

        # Only try to expand more rows when no visible warning row was fixed in this round.
        if progress == 0:
            expanded = click_visible_item_expanders(page, min_y=items_y)
            progress += expanded
            if expanded:
                dbg(f"Expanded rows this round: {expanded}")

        # Fallback sweep: only after repeated no-progress rounds, widen scope beyond warning rows.
        if progress == 0 and attempts_without_progress >= 3:
            extra_shipto, extra_wbs = process_visible_items_with_errors(
                page,
                meta,
                min_y=items_y,
                include_all_visible_items=True,
            )
            filled_shipto += extra_shipto
            filled_wbs += extra_wbs
            progress += extra_shipto + extra_wbs
            if extra_shipto or extra_wbs:
                dbg(
                    "Fallback broad pass changed "
                    f"ShipTo={extra_shipto}, WBS={extra_wbs}."
                )

        # Prioritize current viewport first: warning rows can appear after expanding/filling.
        # Scroll only when a round made no progress.
        try:
            if progress > 0:
                dbg("Progress made; re-checking current viewport before scrolling.")
            else:
                # No progress: use small steps to avoid skipping alternating rows.
                # Periodically jump near top for a fresh pass over virtualized items.
                if attempts_without_progress > 0 and attempts_without_progress % 6 == 0:
                    scroll_delta = -2200
                    page.mouse.wheel(0, scroll_delta)
                    dbg("Jumped to top for full rescan.")
                else:
                    scroll_delta = 320
                    page.mouse.wheel(0, scroll_delta)
                    dbg("Scrolled down slightly to scan adjacent item rows.")
        except PlaywrightError:
            pass

        remaining_after = count_visible_item_error_banners(page, min_y=items_y)
        if remaining_after == 0:
            zero_banner_rounds += 1
        else:
            zero_banner_rounds = 0
        if DEBUG:
            round_elapsed = time.monotonic() - round_start
            print(
                "[DEBUG] round "
                f"{round_idx + 1}/{max_rounds}: "
                f"before={remaining_before}, after={remaining_after}, "
                f"shipto+={progress_shipto}, wbs+={progress_wbs}, "
                f"expanded={expanded}, fallback_shipto+={extra_shipto}, fallback_wbs+={extra_wbs}, "
                f"scroll={scroll_delta}, progress={progress}, "
                f"no_progress_streak={attempts_without_progress}, "
                f"elapsed={round_elapsed:.1f}s"
            )

        # Fast exit: no visible item errors and expected item coverage reached.
        # This avoids long post-completion scans.
        shipto_done = (expected_items <= 0) or (filled_shipto >= expected_items)
        wbs_done = (not meta.wbs) or (expected_items <= 0) or (filled_wbs >= expected_items)
        if zero_banner_rounds >= 1 and shipto_done and wbs_done:
            dbg(
                "Stopping item-field loop early: no visible error banners and "
                f"coverage reached (ShipTo={filled_shipto}/{expected_items}, WBS={filled_wbs}/{expected_items})."
            )
            break

        if progress == 0:
            attempts_without_progress += 1
            dbg("No progress this round.")
        else:
            attempts_without_progress = 0

        if attempts_without_progress >= 22:
            dbg("Stopping item-field loop: repeated no-progress rounds.")
            break

    print(
        "Filled item-level fields on requisition: "
        f"Need-by Date=header-only, Deliver To=header-only, ShipTo(Plant)={filled_shipto}, WBS={filled_wbs}"
    )
    remaining = count_visible_item_error_banners(page, min_y=items_y)
    dbg(
        f"Final item field coverage: "
        f"WBS={filled_wbs}/{expected_items}, remaining-error-banners={remaining}"
    )

    if expected_items > 0 and meta.wbs and filled_wbs < expected_items:
        print(
            f"Warning: could not set item WBS for all items (set {filled_wbs}/{expected_items}). "
            "Please complete remaining item(s) manually."
        )
    if remaining > 0:
        print(
            f"Warning: {remaining} visible item error banner(s) remain. "
            "Please verify before submit."
        )


def count_visible_item_error_banners(page: Page, min_y: float) -> int:
    banners = page.locator("span:has-text('This item contains missing or incorrect information.'):visible")
    try:
        count = banners.count()
    except PlaywrightError:
        return 0
    visible = 0
    for i in range(count):
        b = banners.nth(i)
        try:
            if not b.is_visible():
                continue
            box = b.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            visible += 1
        except PlaywrightError:
            continue
    return visible


def process_visible_items_with_errors(
    page: Page,
    meta: OrderMeta,
    min_y: float,
    include_all_visible_items: bool = False,
) -> tuple[int, int]:
    shipto_changed = 0
    wbs_changed = 0
    # Prefer explicit visible item roots. Banner ancestor chains can be unstable in some Ariba layouts.
    item_containers: List[tuple[float, Locator]] = []
    seen_item_y: List[float] = []
    seen_item_ids: set[str] = set()
    items = page.locator("[aria-label^='Item ']:visible")
    try:
        count = items.count()
    except PlaywrightError:
        count = 0
    for i in range(count):
        item = items.nth(i)
        try:
            if not item.is_visible():
                continue
            box = item.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            if any(abs(box["y"] - y) < 6 for y in seen_item_y):
                continue
            item_text_low = safe_text(item, timeout_ms=900).lower()
            has_error = "this item contains missing or incorrect information" in item_text_low
            if not include_all_visible_items and not has_error:
                continue
            item_id = safe_attr(item, "aria-label")
            if item_id:
                if item_id in seen_item_ids:
                    continue
                seen_item_ids.add(item_id)
            seen_item_y.append(box["y"])
            item_containers.append((box["y"], item))
        except PlaywrightError:
            continue

    # Some Ariba builds do not expose every row with aria-label='Item ...'.
    # Include visible line-item containers too, with the same filtering rules.
    containers = page.locator("div[class*='line-item-container-']:visible")
    try:
        ccount = containers.count()
    except PlaywrightError:
        ccount = 0
    for i in range(ccount):
        item = containers.nth(i)
        try:
            if not item.is_visible():
                continue
            box = item.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            if any(abs(box["y"] - y) < 6 for y in seen_item_y):
                continue
            item_text_low = safe_text(item, timeout_ms=900).lower()
            has_error = "this item contains missing or incorrect information" in item_text_low
            if not include_all_visible_items and not has_error:
                continue
            item_id = safe_attr(item, "aria-label")
            if item_id:
                if item_id in seen_item_ids:
                    continue
                seen_item_ids.add(item_id)
            seen_item_y.append(box["y"])
            item_containers.append((box["y"], item))
        except PlaywrightError:
            continue

    # If no explicit error rows were discovered, use banner Y-position as a fallback anchor.
    if not item_containers and not include_all_visible_items:
        banners = page.locator("span:has-text('This item contains missing or incorrect information.'):visible")
        try:
            bcount = banners.count()
        except PlaywrightError:
            bcount = 0
        for i in range(bcount):
            b = banners.nth(i)
            try:
                if not b.is_visible():
                    continue
                box = b.bounding_box()
                if box is None or box["y"] <= min_y:
                    continue
                nearest = nearest_visible_item_by_y(page, box["y"], min_y=min_y)
                if nearest is None:
                    continue
                nbox = nearest.bounding_box()
                if nbox is None:
                    continue
                item_id = safe_attr(nearest, "aria-label")
                if item_id:
                    if item_id in seen_item_ids:
                        continue
                    seen_item_ids.add(item_id)
                if any(abs(nbox["y"] - y) < 6 for y in seen_item_y):
                    continue
                seen_item_y.append(nbox["y"])
                item_containers.append((nbox["y"], nearest))
            except PlaywrightError:
                continue

    # Process strictly top-to-bottom on screen.
    item_containers.sort(key=lambda t: t[0])
    for item_y, item in item_containers:
        item_label = safe_attr(item, "aria-label")
        dbg(f"Processing item container near y={item_y:.0f}, label='{item_label}'")
        try:
            item.scroll_into_view_if_needed(timeout=1200)
            page.wait_for_timeout(180)
        except PlaywrightError:
            pass

        opened = ensure_item_details_open(item, page, item_y_hint=item_y)
        if not opened:
            # Retry with nearest visible explicit item root for this Y.
            nearest = nearest_visible_item_by_y(page, item_y, min_y=min_y)
            if nearest is not None:
                item = nearest
                item_label = safe_attr(item, "aria-label")
                dbg(f"Retrying open on nearest explicit item near y={item_y:.0f}, label='{item_label}'")
                opened = ensure_item_details_open(item, page, item_y_hint=item_y)
        if opened:
            page.wait_for_timeout(180)
        else:
            dbg(f"Could not confirm item details opened for item near y={item_y:.0f}; continuing scan.")
            continue

        if set_item_shipto_in_scope(page, item):
            shipto_changed += 1
        if meta.wbs and set_item_wbs_in_scope(page, item, meta.wbs):
            wbs_changed += 1

    return shipto_changed, wbs_changed


def ensure_item_details_open(item: Locator, page: Page, item_y_hint: Optional[float] = None) -> bool:
    def is_locator_visible(locator: Locator) -> bool:
        try:
            count = locator.count()
        except PlaywrightError:
            return False
        for i in range(min(count, 4)):
            try:
                if locator.nth(i).is_visible():
                    return True
            except PlaywrightError:
                continue
        return False

    def is_open() -> bool:
        return (
            is_locator_visible(item.locator("span.heading:has-text('Accounting')"))
            or is_locator_visible(item.locator("xpath=.//*[contains(normalize-space(.), 'Account Assignment')]"))
            or is_locator_visible(item.locator("button[role='combobox'][data-help-id*='-AccountCategory']"))
            or is_locator_visible(item.locator("button[role='combobox'][data-help-id*='-WBSElement']"))
            or is_locator_visible(item.locator("button[role='combobox'][data-help-id*='-ShipTo']"))
        )

    if is_open():
        return True

    for _ in range(3):
        # Primary method from recorder.
        try:
            toggle = item.get_by_role(
                "button",
                name=re.compile("Toggle item details|Expand item details|Expand", re.IGNORECASE),
            ).first
            if toggle.count() > 0 and toggle.is_visible():
                toggle.click(timeout=1300)
                page.wait_for_timeout(700)
                if is_open():
                    return True
        except PlaywrightError:
            pass

        # Fallback: click left chevron/arrow inside this item.
        try:
            chevron = item.locator(
                "xpath=.//*[self::button or self::span or self::i]"
                "[contains(@class,'icon-slim-arrow-right') or contains(@class,'icon-slim-arrow-left') "
                "or normalize-space(text())='>' or normalize-space(text())='‹' or contains(@aria-label,'Toggle item details')]"
            ).first
            if chevron.count() > 0 and chevron.is_visible():
                chevron.click(timeout=1300)
                page.wait_for_timeout(1500)
                if is_open():
                    return True
        except PlaywrightError:
            pass

        # Fallback: click the warning text that asks to expand details.
        try:
            warning_expand = item.locator(
                "xpath=.//*[contains(normalize-space(.), 'Expand to review the fields highlighted in red')]"
            ).first
            if warning_expand.count() > 0 and warning_expand.is_visible():
                warning_expand.click(timeout=1200)
                page.wait_for_timeout(700)
                if is_open():
                    return True
        except PlaywrightError:
            pass

        # Last resort: click near the left edge of the item row.
        try:
            box = item.bounding_box()
            if box is not None:
                page.mouse.click(box["x"] + 14, box["y"] + min(70, max(20, box["height"] / 2)))
                page.wait_for_timeout(700)
                if is_open():
                    return True
        except PlaywrightError:
            pass

        # Final fallback: click nearest visible left-side expander around this item Y.
        target_y = item_y_hint
        if target_y is None:
            try:
                box = item.bounding_box()
                target_y = box["y"] if box else None
            except PlaywrightError:
                target_y = None
        if target_y is not None and click_nearest_item_expander_by_y(page, float(target_y)):
            page.wait_for_timeout(750)
            if is_open():
                return True

    return False


def click_nearest_item_expander_by_y(page: Page, target_y: float) -> bool:
    candidates = page.locator(
        "xpath=("
        "//*[self::button or self::a or self::span or self::i]"
        "[contains(@aria-label,'Toggle item details') or contains(@aria-label,'Expand') "
        "or contains(@title,'Expand') or contains(@class,'icon-slim-arrow-right') "
        "or contains(@class,'icon-slim-arrow-left') or contains(@class,'arrow-right') "
        "or contains(@class,'chevron-right') or normalize-space(text())='>']"
        "| //*[@role='button' and (contains(@aria-label,'Toggle') or contains(@aria-label,'Expand') or contains(@title,'Expand'))]"
        ")"
    )
    try:
        count = candidates.count()
    except PlaywrightError:
        return False

    best = None
    best_dist = 10_000.0
    for i in range(min(count, 140)):
        c = candidates.nth(i)
        try:
            if not c.is_visible():
                continue
            box = c.bounding_box()
            if box is None:
                continue
            if box["x"] > 220:
                continue
            dist = abs(box["y"] - target_y)
            if dist > 220:
                continue
            if dist < best_dist:
                best_dist = dist
                best = c
        except PlaywrightError:
            continue

    if best is None:
        return False
    try:
        best.click(timeout=1200)
        return True
    except PlaywrightError:
        return False


def nearest_visible_item_by_y(page: Page, target_y: float, min_y: float = -1.0) -> Optional[Locator]:
    items = page.locator("[aria-label^='Item ']")
    try:
        count = items.count()
    except PlaywrightError:
        return None

    best = None
    best_dist = 10_000.0
    for i in range(min(count, 160)):
        item = items.nth(i)
        try:
            if not item.is_visible():
                continue
            box = item.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            dist = abs(box["y"] - target_y)
            if dist > 260:
                continue
            if dist < best_dist:
                best_dist = dist
                best = item
        except PlaywrightError:
            continue
    return best


def set_item_shipto_in_scope(page: Page, item: Locator) -> bool:
    dbg("Attempting ShipTo set in current item scope.")
    control = item.locator("button[role='combobox'][data-help-id*='-ShipTo']").first
    try:
        if control.count() == 0 or not control.is_visible():
            return False
    except PlaywrightError:
        return False

    current = clean_text(control.text_content())
    if ITEM_SHIPTO_TARGET_PREFIX.lower() in current.lower():
        return False

    try:
        control.click(timeout=1200)
        page.wait_for_timeout(1400)
    except PlaywrightError:
        dbg("ShipTo control click failed.")
        return False

    option = page.locator(
        "a.user-choice[data-help-id*='-ShipTo-']:visible",
        has_text=re.compile(r"0006\s*\(UvA", re.IGNORECASE),
    ).first
    try:
        option.wait_for(timeout=3500)
        page.wait_for_timeout(800)
        option.click(timeout=1800)
        page.wait_for_timeout(1400)
    except PlaywrightError:
        dbg("ShipTo option selection failed.")
        return False

    now = clean_text(control.text_content())
    return ITEM_SHIPTO_TARGET_PREFIX.lower() in now.lower()


def set_item_wbs_in_scope(page: Page, item: Locator, wbs: str) -> bool:
    dbg("Attempting WBS set in current item scope.")
    accounting_button = item.get_by_role("button", name=re.compile("^Accounting collapsed$", re.IGNORECASE)).first
    try:
        if accounting_button.count() > 0 and accounting_button.is_visible():
            accounting_button.click(timeout=1200)
            page.wait_for_timeout(2400)
    except PlaywrightError:
        pass

    account_control = item.get_by_role("combobox", name=re.compile("Account Assignment", re.IGNORECASE)).first
    try:
        if account_control.count() == 0 or not account_control.is_visible():
            dbg("Account Assignment control not visible in this item scope.")
            return False
    except PlaywrightError:
        return False

    assignment_now = safe_text(account_control)
    if "P (Project)" not in assignment_now:
        try:
            account_control.click(timeout=1200)
            page.wait_for_timeout(1300)
            # Prefer the explicit AccountCategory option in the currently open chooser.
            chooser = account_control.locator(
                "xpath=ancestor::*[contains(@class,'field-chooser')][1]"
            ).first
            p_option = chooser.locator(
                "xpath=.//a[contains(@class,'user-choice') and contains(@data-help-id,'-AccountCategory-') and contains(normalize-space(.),'P (Project)')]"
            ).first
            if p_option.count() == 0:
                p_option = page.locator(
                    "a.user-choice[data-help-id*='-AccountCategory-0']:visible"
                ).first
            if p_option.count() == 0:
                p_option = page.locator(
                    "a.user-choice[data-help-id*='-AccountCategory-']:has-text('P (Project)'):visible"
                ).first
            p_option.wait_for(timeout=4500)
            page.wait_for_timeout(700)
            try:
                p_option.click(timeout=2200)
            except PlaywrightError:
                # Ariba dropdowns occasionally intercept pointer events briefly.
                p_option.evaluate("el => el.click()")
            page.wait_for_timeout(1500)
        except PlaywrightError:
            dbg("Selecting Account Assignment 'P (Project)' failed.")
            return False
        except Exception:
            return False
        assignment_now = safe_text(account_control)
        if "P (Project)" not in assignment_now:
            return False

    wbs_control = item.locator("button[role='combobox'][data-help-id*='-WBSElement']").first
    if wbs_control.count() == 0:
        wbs_control = item.get_by_role("combobox", name=re.compile("Project/WBS", re.IGNORECASE)).first
    try:
        if wbs_control.count() == 0 or not wbs_control.is_visible():
            dbg("WBS control not visible in this item scope.")
            return False
    except PlaywrightError:
        return False

    if wbs.lower() in safe_text(wbs_control).lower():
        return True

    try:
        # Open WBSElement dropdown (single click + wait for slow UI).
        wbs_control.click(timeout=1300)
        page.wait_for_timeout(1400)
    except PlaywrightError:
        try:
            wbs_control.evaluate("el => el.click()")
            page.wait_for_timeout(1400)
        except Exception:
            dbg("WBS dropdown open failed.")
            return False

    # Scope option lookup to same chooser when possible.
    chooser = wbs_control.locator("xpath=ancestor::*[contains(@class,'field-chooser')][1]").first
    # Prefer Browse all because WBS varies often and may not be listed directly.
    try:
        browse = chooser.locator("a.browse-all[data-help-id*='-WBSElement']").first
        if browse.count() == 0:
            browse = page.locator("a.browse-all[data-help-id*='-WBSElement']:visible").first
        browse.wait_for(timeout=3500)
        page.wait_for_timeout(700)
        browse.click(timeout=1800)
        page.wait_for_timeout(1400)
        search = page.locator("input[data-help-id*='-WBSElement'][placeholder='Search']:visible").first
        if search.count() == 0:
            search = page.get_by_role("textbox", name=re.compile("Type search term here", re.IGNORECASE)).locator(":visible").first
        search.wait_for(timeout=5000)
        search.click(timeout=1200)
        search.fill(wbs)
        page.wait_for_timeout(1100)
        search.press("Enter")
        page.wait_for_timeout(1000)
        search_btn = page.get_by_role("button", name=re.compile("^Search$", re.IGNORECASE)).locator(":visible").first
        if search_btn.count() > 0:
            search_btn.click(timeout=1500)
        page.wait_for_timeout(1500)
        choose_btn = page.locator(
            "button.chooser-choose-button[data-help-id*='-WBSElement-']:visible"
        ).first
        if choose_btn.count() == 0:
            choose_btn = page.get_by_role(
                "button",
                name=re.compile(rf"choose Project/WBS {re.escape(wbs)}", re.IGNORECASE),
            ).locator(":visible").first
        choose_btn.wait_for(timeout=6000)
        choose_btn.click(timeout=2200)
        page.wait_for_timeout(2300)
    except PlaywrightError:
        dbg("WBS modal/chooser path failed; trying direct option fallback.")
        # Fallback: direct option in open dropdown if modal path fails.
        direct_option = chooser.locator(
            "a.user-choice[data-help-id*='-WBSElement-']",
            has_text=re.compile(re.escape(wbs), re.IGNORECASE),
        ).first
        if direct_option.count() == 0:
            direct_option = page.locator(
                "a.user-choice[data-help-id*='-WBSElement-']:visible",
                has_text=re.compile(re.escape(wbs), re.IGNORECASE),
            ).first
        try:
            direct_option.wait_for(timeout=3000)
            page.wait_for_timeout(700)
            direct_option.click(timeout=1800)
            page.wait_for_timeout(1800)
        except PlaywrightError:
            dbg("Direct WBS option fallback failed.")
            return False

    final_wbs_text = safe_text(wbs_control).lower()
    if wbs.lower() in final_wbs_text:
        return True
    if "(no value)" not in final_wbs_text and final_wbs_text:
        # Accept any non-empty chosen Project/WBS value.
        return True
    # Some Ariba views lag combobox label updates; confirm using full item text as fallback.
    return wbs.lower() in safe_text(item, timeout_ms=1200).lower()


def set_item_shipto_dropdowns(page: Page, min_y: float) -> int:
    changed = 0
    shipto_labels = page.locator("xpath=//*[starts-with(normalize-space(.), 'ShipTo(Plant)')]")
    label_count = shipto_labels.count()

    for i in range(label_count):
        label = shipto_labels.nth(i)
        try:
            if not label.is_visible():
                continue
            label_box = label.bounding_box()
            if label_box is None or label_box["y"] <= min_y:
                continue
        except PlaywrightError:
            continue

        # ShipTo controls are typically in the same row, to the right of label, and close in Y.
        row_container = None
        try:
            row_container = label.locator(
                "xpath=ancestor::*[contains(@class,'ship-section') or contains(@class,'line-item-sections') or contains(@class,'line-item-container')][1]"
            ).first
            if row_container.count() == 0:
                row_container = label.locator("xpath=ancestor::*[self::div or self::section][1]").first
        except PlaywrightError:
            row_container = None

        control = first_visible_or_none(
            [
                row_container.locator("xpath=.//button[@role='combobox' and contains(@data-help-id,'-ShipTo')]"),
                label.locator("xpath=following::button[@role='combobox' and contains(@data-help-id,'-ShipTo')][position()<=4]"),
                page.locator("button[role='combobox'][data-help-id*='-ShipTo']:visible"),
                label.locator("xpath=following::*[@role='combobox'][position()<=8]"),
                label.locator("xpath=following::input[position()<=8]"),
                label.locator("xpath=following::button[position()<=8]"),
                label.locator("xpath=following::*[contains(@class,'select') or contains(@class,'dropdown')][position()<=8]"),
            ]
        )
        if control is None:
            continue
        try:
            if not control.is_visible():
                continue
            box = control.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            # keep control near label and to its right
            if abs(box["y"] - label_box["y"]) > 120:
                continue
            if box["x"] + 20 < label_box["x"]:
                continue
        except PlaywrightError:
            continue

        current_value = ""
        try:
            current_value = clean_text(control.input_value())
        except PlaywrightError:
            try:
                current_value = clean_text(control.text_content())
            except PlaywrightError:
                current_value = ""

        row_text = ""
        if row_container is not None:
            try:
                row_text = clean_text(row_container.text_content())
            except PlaywrightError:
                row_text = ""

        if (
            ITEM_SHIPTO_TARGET_PREFIX.lower() in current_value.lower()
            or ITEM_SHIPTO_TARGET_PREFIX.lower() in row_text.lower()
        ):
            continue
        dbg(f"ShipTo current value before change: '{current_value}'")

        opened = False
        try:
            control.click(timeout=1000)
            page.wait_for_timeout(1200)
            if page.locator("a.user-choice[data-help-id*='-ShipTo-0']:visible").count() > 0:
                opened = True
        except PlaywrightError:
            pass
        if not opened:
            try:
                label.click(timeout=900)
                page.wait_for_timeout(1200)
                if page.locator("a.user-choice[data-help-id*='-ShipTo-0']:visible").count() > 0:
                    opened = True
            except PlaywrightError:
                pass
        if not opened:
            dbg("ShipTo dropdown did not open after retries.")
            continue

        field_chooser = None
        try:
            field_chooser = control.locator("xpath=ancestor::*[contains(@class,'field-chooser')][1]").first
        except PlaywrightError:
            field_chooser = None

        clicked_option = False
        option_candidates = [
            field_chooser.locator(
                "xpath=.//a[contains(@class,'user-choice') and contains(@data-help-id,'-ShipTo-') and contains(normalize-space(.),'0006')]"
            ).first if field_chooser is not None else page.locator("xpath=//never").first,
            field_chooser.locator(
                "xpath=.//a[contains(@class,'user-choice') and contains(@data-help-id,'-ShipTo-') and contains(normalize-space(.),'UvA')]"
            ).first if field_chooser is not None else page.locator("xpath=//never").first,
            page.locator("a.user-choice[data-help-id*='-ShipTo-0']:visible").first,
            page.locator("a.user-choice[data-help-id*='-ShipTo-']:has-text('0006'):visible").first,
            page.locator(f"xpath=//*[contains(normalize-space(.), '{ITEM_SHIPTO_TARGET_PREFIX}')]").first,
        ]
        for option in option_candidates:
            if option.count() == 0:
                continue
            try:
                option.wait_for(timeout=3000)
                page.wait_for_timeout(900)
                option.click(timeout=1500)
                page.wait_for_timeout(1200)
                clicked_option = True
                break
            except PlaywrightError:
                continue

        if not clicked_option:
            # Fallback: keyboard-select first option after opening.
            try:
                control.click(timeout=400)
                page.wait_for_timeout(700)
                control.press("ArrowDown")
                page.wait_for_timeout(300)
                control.press("Enter")
                page.wait_for_timeout(1200)
                clicked_option = True
            except PlaywrightError:
                dbg("ShipTo option click/navigation failed for visible control.")
                continue

        # Confirm actual change to target value.
        deadline = time.monotonic() + 2.0
        confirmed = False
        while time.monotonic() < deadline:
            try:
                now = clean_text(control.input_value())
            except PlaywrightError:
                now = clean_text(control.text_content())
            row_now = ""
            if row_container is not None:
                try:
                    row_now = clean_text(row_container.text_content())
                except PlaywrightError:
                    row_now = ""
            if (
                ITEM_SHIPTO_TARGET_PREFIX.lower() in now.lower()
                or ITEM_SHIPTO_TARGET_PREFIX.lower() in row_now.lower()
            ):
                confirmed = True
                break
            page.wait_for_timeout(150)
        if confirmed:
            changed += 1
            dbg("ShipTo changed successfully.")
        else:
            dbg("ShipTo change could not be confirmed.")

    return changed


def set_item_wbs_fields(page: Page, wbs: str, min_y: float) -> int:
    changed = 0
    # Ensure accounting sections are expanded for visible rows.
    expand_visible_accounting_sections(page, min_y=min_y)

    account_controls = page.locator("button[role='combobox'][data-help-id*='-AccountCategory']:visible")
    try:
        count = account_controls.count()
    except PlaywrightError:
        return 0

    for i in range(count):
        account_control = account_controls.nth(i)
        try:
            if not account_control.is_visible():
                continue
            box = account_control.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
        except PlaywrightError:
            continue

        # Work within the same item block.
        row_container = account_control.locator(
            "xpath=ancestor::*[self::div or self::section][1]"
        ).first

        current_assignment = clean_text(account_control.text_content())
        if "P (Project)" not in current_assignment:
            set_project_ok = False
            for _ in range(2):
                try:
                    account_control.click(timeout=1000)
                    # Ariba is slow to render dropdown choices.
                    page.wait_for_timeout(1300)
                    field_chooser = account_control.locator(
                        "xpath=ancestor::*[contains(@class,'field-chooser')][1]"
                    ).first
                    option = field_chooser.locator(
                        "xpath=.//a[contains(@class,'user-choice') and contains(@data-help-id,'-AccountCategory-') and contains(normalize-space(.),'P (Project)')]"
                    ).first
                    if option.count() == 0:
                        option = page.locator(
                            "a.user-choice[data-help-id*='-AccountCategory-']:has-text('P (Project)'):visible"
                        ).first
                    option.wait_for(timeout=4000)
                    page.wait_for_timeout(700)
                    option.click(timeout=1800)
                    # Let the selection commit and WBSElement appear.
                    page.wait_for_timeout(1500)
                    now_assignment = clean_text(account_control.text_content())
                    if "P (Project)" in now_assignment:
                        set_project_ok = True
                        break
                except PlaywrightError:
                    try:
                        account_control.press("p")
                        page.wait_for_timeout(300)
                        account_control.press("Enter")
                        page.wait_for_timeout(1500)
                        now_assignment = clean_text(account_control.text_content())
                        if "P (Project)" in now_assignment:
                            set_project_ok = True
                            break
                    except PlaywrightError:
                        continue
            if not set_project_ok:
                dbg("Could not set Account Assignment to P (Project) for one row.")
                continue
            dbg("Set Account Assignment to P (Project) for one row.")

        # Now set Project/WBS in WBSElement chooser.
        cost_control = row_container.locator(
            "xpath=.//button[@role='combobox' and contains(@data-help-id,'-WBSElement')]"
        ).first
        if cost_control.count() == 0:
            cost_control = row_container.locator(
                "xpath=.//button[@role='combobox' and contains(@data-help-id,'-CostCenter')]"
            ).first
        if cost_control.count() == 0:
            cost_control = page.locator(
                "button[role='combobox'][data-help-id*='-WBSElement']:visible"
            ).first
        if cost_control.count() == 0:
            cost_control = page.locator(
                "button[role='combobox'][data-help-id*='-CostCenter']:visible"
            ).first
        try:
            if cost_control.count() == 0 or not cost_control.is_visible():
                continue
        except PlaywrightError:
            continue

        current_cost = clean_text(cost_control.text_content())
        row_text = ""
        try:
            row_text = clean_text(row_container.text_content())
        except PlaywrightError:
            row_text = ""

        if wbs.lower() in current_cost.lower() or wbs.lower() in row_text.lower():
            continue

        if set_one_costcenter_wbs(page, cost_control, wbs):
            changed += 1

    return changed


def expand_visible_accounting_sections(page: Page, min_y: float) -> int:
    expanded = 0
    # In this tenant Accounting toggle is a button with aria-label like "Accounting collapsed".
    # Target only buttons (not inner spans) to avoid double-toggle closing.
    headings = page.locator(
        "button.btn-link[aria-label*='Accounting'], "
        "button[aria-label*='Accounting']"
    )
    try:
        count = headings.count()
    except PlaywrightError:
        return 0
    seen_y: List[float] = []

    for i in range(count):
        heading = headings.nth(i)
        try:
            if not heading.is_visible():
                continue
            hbox = heading.bounding_box()
            if hbox is None or hbox["y"] <= min_y:
                continue
            if any(abs(hbox["y"] - y) < 6 for y in seen_y):
                continue
            seen_y.append(hbox["y"])
        except PlaywrightError:
            continue

        # Resolve the actual toggle button for this heading.
        try:
            aria = clean_text(heading.get_attribute("aria-label")).lower()
            if "accounting" in aria:
                toggle_btn = heading
            else:
                toggle_btn = heading.locator(
                    "xpath=ancestor::button[contains(@aria-label,'Accounting') or contains(@aria-label,'accounting') or contains(@class,'btn-link')][1]"
                ).first
        except PlaywrightError:
            continue

        # Click only when collapsed.
        try:
            aria_now = clean_text(toggle_btn.get_attribute("aria-label")).lower()
            collapsed_marker = toggle_btn.locator(".icon-slim-arrow-right").first
            is_collapsed = ("collapsed" in aria_now) or (collapsed_marker.count() > 0)
            if not is_collapsed:
                continue
        except PlaywrightError:
            continue

        opened = False
        try:
            toggle_btn.click(timeout=1000)
            # Single click then wait; second click would collapse it again.
            page.wait_for_timeout(2600)
            aria_after = clean_text(toggle_btn.get_attribute("aria-label")).lower()
            marker_after = toggle_btn.locator(".icon-slim-arrow-right").count()
            # Consider opened if collapsed indicator disappears or aria no longer says collapsed.
            if ("collapsed" not in aria_after) and marker_after == 0:
                opened = True
            else:
                # Some builds keep aria stale; verify by visibility of AccountCategory in same item area.
                row_scope = toggle_btn.locator(
                    "xpath=ancestor::*[contains(@class,'line-item') or contains(@class,'core-section') or self::div][1]"
                ).first
                account_control = row_scope.locator(
                    "xpath=.//button[contains(@data-help-id,'-AccountCategory')]"
                ).first
                if account_control.count() > 0 and account_control.is_visible():
                    opened = True
        except PlaywrightError:
            opened = False
        if opened:
            expanded += 1
            dbg("Opened one Accounting panel.")
        else:
            dbg("Could not confirm Accounting panel opening for one visible row.")
    if expanded:
        dbg(f"Expanded accounting sections this pass: {expanded}")
    return expanded


def count_item_needby_matches(page: Page, value: str, min_y: float) -> int:
    selectors = [
        "input[data-help-id*='NeedBy']:visible",
        "input[data-help-id*='Need-by']:visible",
        "input[aria-label*='Need-by Date' i]:visible",
        "input[name='NeedByDate']:visible",
    ]
    count_matches = 0
    seen_y = []
    for selector in selectors:
        loc = page.locator(selector)
        try:
            count = loc.count()
        except PlaywrightError:
            continue
        for i in range(count):
            inp = loc.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None or box["y"] <= min_y:
                    continue
                if any(abs(box["y"] - y) < 6 for y in seen_y):
                    continue
                now = clean_text(inp.input_value())
                if now == clean_text(value):
                    count_matches += 1
                seen_y.append(box["y"])
            except PlaywrightError:
                continue
    return count_matches


def count_item_deliver_matches(page: Page, value: str, min_y: float) -> int:
    selectors = [
        "input[name='DeliverTo'][data-help-id*='-DeliverTo']:visible",
        "input[name='DeliverTo']:visible",
        "input[aria-label*='Deliver To' i]:visible",
    ]
    count_matches = 0
    seen_y = []
    for selector in selectors:
        loc = page.locator(selector)
        try:
            count = loc.count()
        except PlaywrightError:
            continue
        for i in range(count):
            inp = loc.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None or box["y"] <= min_y:
                    continue
                if any(abs(box["y"] - y) < 6 for y in seen_y):
                    continue
                now = clean_text(inp.input_value())
                if now == clean_text(value):
                    count_matches += 1
                seen_y.append(box["y"])
            except PlaywrightError:
                continue
    return count_matches


def count_item_shipto_matches(page: Page, min_y: float) -> int:
    controls = page.locator("button[role='combobox'][data-help-id*='-ShipTo']:visible")
    try:
        count = controls.count()
    except PlaywrightError:
        return 0
    matches = 0
    for i in range(count):
        c = controls.nth(i)
        try:
            if not c.is_visible():
                continue
            box = c.bounding_box()
            if box is None or box["y"] <= min_y:
                continue
            txt = clean_text(c.text_content())
            if ITEM_SHIPTO_TARGET_PREFIX.lower() in txt.lower():
                matches += 1
        except PlaywrightError:
            continue
    return matches


def set_one_costcenter_wbs(page: Page, cost_control, wbs: str) -> bool:
    # Open Project/WBS chooser (menu can appear with delay).
    clicked_browse = False
    for _ in range(2):
        try:
            cost_control.click(timeout=1000)
        except PlaywrightError:
            continue
        page.wait_for_timeout(1200)

        browse_all = page.locator("a.browse-all[data-help-id*='-WBSElement']:visible").first
        if browse_all.count() == 0:
            browse_all = page.locator("a.browse-all[data-help-id*='-CostCenter']:visible").first
        if browse_all.count() == 0:
            continue
        try:
            browse_all.click(timeout=1200)
            page.wait_for_timeout(1200)
            clicked_browse = True
            break
        except PlaywrightError:
            continue

    if not clicked_browse:
        # Fallback: type directly into control.
        try:
            cost_control.click(timeout=400)
            cost_control.type(wbs, delay=20)
            cost_control.press("Enter")
        except PlaywrightError:
            return False
    else:
        # In modal chooser: search and choose.
        search_input = page.locator("input[data-help-id*='-WBSElement'][placeholder='Search']:visible").first
        if search_input.count() == 0:
            search_input = page.locator("input[data-help-id*='-CostCenter'][placeholder='Search']:visible").first
        try:
            search_input.wait_for(timeout=5000)
            search_input.fill("")
            search_input.fill(wbs)
            page.wait_for_timeout(1100)
            search_input.press("Enter")
            page.wait_for_timeout(1800)
        except PlaywrightError:
            return False

        # Prefer explicit "Choose" button pattern from WBSElement chooser.
        choose_candidates = [
            page.locator("button.chooser-choose-button[data-help-id*='-WBSElement-']:visible").first,
            page.locator("button[data-help-id*='-WBSElement-0']:visible").first,
            page.locator("button.chooser-choose-button[data-help-id*='-CostCenter-']:visible").first,
            page.get_by_role("button", name=re.compile(r"^Choose$", re.IGNORECASE)).locator(":visible").first,
        ]
        clicked_choose = False
        for choose_btn in choose_candidates:
            if choose_btn.count() == 0:
                continue
            try:
                # If multiple choices exist, select row by wbs text first when possible.
                matched_cell = page.locator(
                    "tr:visible, [role='row']:visible, td:visible, [role='cell']:visible",
                    has_text=re.compile(re.escape(wbs), re.IGNORECASE),
                ).first
                if matched_cell.count() > 0:
                    try:
                        matched_cell.click(timeout=1200)
                    except PlaywrightError:
                        pass
                choose_btn.wait_for(timeout=5000)
                choose_btn.click(timeout=2000)
                page.wait_for_timeout(2200)
                clicked_choose = True
                break
            except PlaywrightError:
                continue
        if not clicked_choose:
            return False

    # Confirm wbs text is now present in control or nearby item row.
    deadline = time.monotonic() + 6.0
    while time.monotonic() < deadline:
        try:
            now = clean_text(cost_control.text_content())
            if wbs.lower() in now.lower():
                return True
        except PlaywrightError:
            pass
        page.wait_for_timeout(150)
    dbg("WBS selection could not be confirmed.")
    return False


def fill_item_needby_inputs(page: Page, value: str, min_y: float) -> int:
    selectors = [
        "input[data-help-id*='NeedBy']:visible",
        "input[data-help-id*='Need-by']:visible",
        "input[aria-label*='Need-by Date' i]:visible",
        "input[name='NeedByDate']:visible",
    ]
    changed = 0
    seen_y = []
    for selector in selectors:
        loc = page.locator(selector)
        try:
            count = loc.count()
        except PlaywrightError:
            continue
        for i in range(count):
            inp = loc.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None or box["y"] <= min_y:
                    continue
                if any(abs(box["y"] - y) < 6 for y in seen_y):
                    continue
                was_changed = fill_input_if_needed(inp, value, page)
                if was_changed:
                    changed += 1
                seen_y.append(box["y"])
            except PlaywrightError:
                continue
    return changed


def fill_item_deliver_inputs(page: Page, value: str, min_y: float) -> int:
    selectors = [
        "input[name='DeliverTo'][data-help-id*='-DeliverTo']:visible",
        "input[name='DeliverTo']:visible",
        "input[aria-label*='Deliver To' i]:visible",
    ]
    changed = 0
    seen_y = []
    for selector in selectors:
        loc = page.locator(selector)
        try:
            count = loc.count()
        except PlaywrightError:
            continue
        for i in range(count):
            inp = loc.nth(i)
            try:
                if not inp.is_visible():
                    continue
                box = inp.bounding_box()
                if box is None or box["y"] <= min_y:
                    continue
                if any(abs(box["y"] - y) < 6 for y in seen_y):
                    continue
                was_changed = fill_input_if_needed(inp, value, page)
                if was_changed:
                    changed += 1
                seen_y.append(box["y"])
            except PlaywrightError:
                continue
    return changed


def get_items_section_min_y(page: Page) -> float:
    # Primary anchor.
    items_header = first_visible_text(page, "Items (")
    if items_header is not None:
        try:
            box = items_header.bounding_box()
            if box:
                return box["y"]
        except PlaywrightError:
            pass

    # Fallback anchor within items section.
    item_error = first_visible_text(page, "This item contains missing or incorrect information")
    if item_error is not None:
        try:
            box = item_error.bounding_box()
            if box:
                return box["y"] - 40
        except PlaywrightError:
            pass
    return -1.0


def checkout_and_fill_requisition(page: Page, meta: OrderMeta, expected_items: int) -> None:
    click_checkout(page)
    fill_requisition_header_fields(page, meta)
    fill_item_fields_on_requisition(page, meta, expected_items)


def fill_one_item(page: Page, item: ItemRow) -> None:
    wait_for_non_catalog_page(page)
    select_fixed_category(page)
    fill_product_name(page, item.product_name)
    fill_description(page, item.description)
    fill_quantity(page, item.quantity)
    fill_unit_price(page, item.unit_price)


def ensure_supplier_selected(page: Page, supplier_name: str) -> None:
    if is_supplier_already_selected(page, supplier_name):
        return
    select_supplier(page, supplier_name)
    verify_supplier_selected(page, supplier_name)


def run_import(page: Page, meta: OrderMeta, items: List[ItemRow]) -> None:
    wait_for_non_catalog_page(page)

    total = len(items)
    for index, item in enumerate(items, start=1):
        print(f"[{index}/{total}] Filling item {index} into Ariba form...")
        ensure_supplier_selected(page, meta.supplier_name)
        fill_one_item(page, item)
        click_add_to_cart(page)
        print(f"[{index}/{total}] Added item {index} to Ariba cart.")

        if index < total:
            create_new_item(page)
    print("Opening checkout and filling requisition fields...")
    checkout_and_fill_requisition(page, meta, total)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Import non-catalog items from Excel into SAP Ariba."
    )
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to .xlsx file with Supplier name in A1/B1 and headers on row 3.",
    )
    parser.add_argument(
        "--cdp-url",
        default="http://127.0.0.1:9222",
        help="Chrome/Edge remote-debugging endpoint (default: http://127.0.0.1:9222).",
    )
    parser.add_argument(
        "--page-contains",
        default="ariba",
        help="Substring used to choose the target browser tab (default: ariba).",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print detailed debug markers for field-filling and loop progress.",
    )
    parser.add_argument(
        "--payment-shipping",
        default=str(default_payment_shipping_path(script_dir)),
        help=(
            "Path to PaymentAndShipping.xlsx "
            "(default: next to this script, else one folder up)."
        ),
    )
    return parser.parse_args()


def pick_target_page(context_page_list: List[Page], needle: str) -> Optional[Page]:
    needle_lower = needle.lower()
    for page in context_page_list:
        try:
            title = page.title().lower()
            url = page.url.lower()
        except PlaywrightError:
            continue
        if needle_lower in title or needle_lower in url:
            return page
    return context_page_list[0] if context_page_list else None


def main() -> int:
    global DEBUG
    args = parse_args()
    DEBUG = args.debug
    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        return 2
    payment_shipping_path = Path(args.payment_shipping).expanduser().resolve()
    if not payment_shipping_path.exists():
        print(f"Payment/shipping file not found: {payment_shipping_path}")
        return 2

    try:
        meta_from_order, items = load_items_from_excel(excel_path)
        meta = load_payment_and_shipping_meta(
            payment_shipping_path, supplier_name=meta_from_order.supplier_name
        )
    except Exception as exc:  # pylint: disable=broad-except
        print(f"Excel validation error: {exc}")
        return 2

    print(f"Loaded {len(items)} item(s). Supplier: {meta.supplier_name}")
    dbg(
        f"Loaded metadata: need_by_date='{meta.need_by_date}', "
        f"deliver_to='{meta.deliver_to}', wbs='{meta.wbs}'"
    )
    print("Connecting to browser via CDP...")

    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(args.cdp_url)
            all_pages: List[Page] = []
            for context in browser.contexts:
                all_pages.extend(context.pages)
            page = pick_target_page(all_pages, args.page_contains)
            if page is None:
                raise RuntimeError("No open browser tabs found in the CDP session.")
            page.bring_to_front()
            run_import(page, meta, items)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"Automation failed: {exc}")
        return 1

    print("Import completed. Please review cart and continue checkout manually.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
