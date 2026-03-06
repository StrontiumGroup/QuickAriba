#!/usr/bin/env python3
"""
Convert a Schaefter + Kirchhoff offer PDF into the Excel format expected by
ariba_excel_import.py.

Mapping:
- Product name <- "<Part Number> | <Description flattened with ', '>" (max 80 chars)
- Description <- up to 3 description lines below part number (with line breaks)
- Quantity <- "pcs."
- Unit price <- "Unit price" converted from gross to net using VATRate.xlsx[A2]
- Supplier Part Number <- part number from first line of each item
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook


@dataclass
class OfferItem:
    position: str
    supplier_part_number: str
    quantity: str
    unit_price_gross: str
    description_lines: List[str] = field(default_factory=list)

    @property
    def description(self) -> str:
        return "\n".join(self.description_lines)


ROW_RE = re.compile(
    r"^\s*(?P<pos>\d+)\s+(?P<part>\S+)\s+(?P<qty>\d+)\s+(?P<unit>[0-9]+(?:[.,][0-9]+)?)\s+[0-9.,]+\s*EUR\b"
)


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def clean_decimal_text(text: str) -> str:
    value = re.sub(r"[^0-9,.\-]", "", (text or "").strip()).replace(",", ".")
    return value


def build_product_name(supplier_part_number: str, description: str, max_len: int = 80) -> str:
    flattened_desc = ", ".join(part.strip() for part in description.splitlines() if part.strip())
    combined = f"{supplier_part_number} | {flattened_desc}".strip()
    return combined[:max_len]


def load_vat_rate_percent(vat_xlsx_path: Path) -> Decimal:
    try:
        workbook = load_workbook(vat_xlsx_path, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Unable to read VAT rate file: {vat_xlsx_path}") from exc

    raw_value = workbook.active["A2"].value
    if raw_value is None:
        raise RuntimeError(f"VAT rate cell A2 is empty in: {vat_xlsx_path}")

    normalized = str(raw_value).strip().replace("%", "").replace(",", ".")
    try:
        vat_rate = Decimal(normalized)
    except InvalidOperation as exc:
        raise RuntimeError(
            f"VAT rate in A2 is not a valid number: {raw_value!r}"
        ) from exc

    if vat_rate < 0:
        raise RuntimeError(f"VAT rate in A2 must be >= 0, got: {vat_rate}")
    return vat_rate


def gross_to_net_price(gross_price: str, vat_rate_percent: Decimal) -> str:
    gross = Decimal(gross_price)
    divisor = Decimal("1") + (vat_rate_percent / Decimal("100"))
    net = (gross / divisor).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    net_text = format(net, "f").rstrip("0").rstrip(".")
    return net_text or "0"


def extract_text_with_pypdf(pdf_path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return ""

    reader = PdfReader(str(pdf_path))
    chunks: List[str] = []
    for page in reader.pages:
        try:
            text = page.extract_text(extraction_mode="layout")
        except TypeError:
            text = page.extract_text()
        chunks.append(text or "")
    return "\n".join(chunks)


def extract_text_with_pdftotext(pdf_path: Path) -> str:
    tool = find_pdftotext_tool()
    if not tool:
        return ""
    try:
        result = subprocess.run(
            [tool, "-layout", "-enc", "UTF-8", str(pdf_path), "-"],
            check=False,
            capture_output=True,
            text=True,
        )
        return result.stdout or ""
    except Exception:
        return ""


def find_pdftotext_tool() -> str:
    found = shutil.which("pdftotext")
    if found:
        return found

    candidates = [
        Path(r"C:\Program Files\MiKTeX\miktex\bin\x64\pdftotext.exe"),
        Path(r"C:\Program Files (x86)\MiKTeX\miktex\bin\pdftotext.exe"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return ""


def is_ignorable_line(line: str) -> bool:
    lower = line.lower().strip()
    if not lower:
        return True

    ignored_substrings = (
        "schäfter+kirchhoff",
        "schäfter + kirchhoff",
        "kieler straße",
        "kieler str.",
        "our reference:",
        "account:",
        "date:",
        "quotation",
        "dear sir or madam",
        "we thank you for your inquiry",
        "pos item / description",
        "pcs. unit price",
        "total eur",
        "page 1/3",
        "page 2/3",
        "page 3/3",
        "page 1",
        "page 2",
        "page 3",
        "subject to prior sale",
        "sales terms and delivery conditions",
        "data protection notes",
        "please specify the quotation number",
        "we look forward to hearing from you",
        "best regards",
        "optics, metrology, photonics",
        "payment ..............:",
        "terms of delivery ....:",
        "delivery time approx. :",
        "quote valid until ....:",
        "delivery mode ........:",
        "total goods ......:",
        "shipping fee .....:",
        "net value ........:",
        "total ............:",
        "incoterms",
        "sukhamburg.com",
        "fon:",
        "fax:",
        "e-mail:",
        "log4cxx:",
        "miktex requires windows",
    )
    return any(token in lower for token in ignored_substrings)


def parse_items_from_text(text: str) -> List[OfferItem]:
    items: List[OfferItem] = []
    in_table = False
    current_item: OfferItem | None = None

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        lower = line.lower()

        if "pos item / description" in lower and "pcs." in lower and "unit price" in lower:
            in_table = True
            continue

        if not in_table:
            continue

        if "total goods ......:" in lower:
            break

        row_match = ROW_RE.match(line)
        if row_match:
            current_item = OfferItem(
                position=row_match.group("pos"),
                supplier_part_number=clean_text(row_match.group("part")),
                quantity=clean_text(row_match.group("qty")),
                unit_price_gross=clean_decimal_text(row_match.group("unit")),
            )
            items.append(current_item)
            continue

        if current_item is None:
            continue

        if is_ignorable_line(line):
            continue

        normalized = clean_text(line)
        if not normalized:
            continue

        # Ignore rebate lines shown below unit price (e.g. "-3.00%").
        if re.fullmatch(r"-?\d+(?:[.,]\d+)?\s*%", normalized):
            continue

        if len(current_item.description_lines) < 3:
            current_item.description_lines.append(normalized)

    # Keep only rows that have required values.
    return [
        item
        for item in items
        if item.supplier_part_number
        and item.quantity
        and item.unit_price_gross
        and item.description_lines
    ]


def write_output_excel(output_path: Path, supplier_name: str, items: List[OfferItem]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["A2"] = supplier_name

    ws["A4"] = "Product name"
    ws["B4"] = "Description"
    ws["C4"] = "Quantity"
    ws["D4"] = "Unit price"
    ws["E4"] = "Supplier Part Number"

    row = 5
    for item in items:
        description = item.description
        ws.cell(
            row=row,
            column=1,
            value=build_product_name(item.supplier_part_number, description),
        )
        ws.cell(row=row, column=2, value=description)
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=item.unit_price_gross)
        ws.cell(row=row, column=5, value=item.supplier_part_number)
        row += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    scripts_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Convert a Schaefter + Kirchhoff offer PDF into Ariba import Excel format."
    )
    parser.add_argument("--pdf", required=True, help="Path to Schaefter + Kirchhoff offer PDF.")
    parser.add_argument(
        "--out",
        required=True,
        help="Path to output .xlsx file (for ariba_excel_import.py).",
    )
    parser.add_argument(
        "--supplier",
        default="Schäfter + Kirchhoff",
        help="Supplier name written to cell A2 (default: Schäfter + Kirchhoff).",
    )
    parser.add_argument(
        "--vat-xlsx",
        default=str(scripts_dir / "VATRate.xlsx"),
        help="Path to VATRate.xlsx (reads VAT rate from cell A2).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pdf_path = Path(args.pdf).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    vat_xlsx_path = Path(args.vat_xlsx).expanduser().resolve()

    if not pdf_path.exists():
        print(f"Input PDF not found: {pdf_path}")
        return 2

    text = extract_text_with_pypdf(pdf_path)
    if not text.strip():
        text = extract_text_with_pdftotext(pdf_path)

    if not text.strip():
        print(
            "Could not extract text from PDF. Install pypdf or ensure pdftotext is available."
        )
        return 2

    items = parse_items_from_text(text)
    if not items:
        print("No offer rows found in PDF table.")
        return 1

    vat_rate_percent = load_vat_rate_percent(vat_xlsx_path)
    for item in items:
        item.unit_price_gross = gross_to_net_price(item.unit_price_gross, vat_rate_percent)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(out_path, args.supplier, items)
    print(f"Wrote {len(items)} item(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
