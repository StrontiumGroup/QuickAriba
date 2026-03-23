#!/usr/bin/env python3
"""
Convert an RS Components cart CSV into the Excel format expected by ariba_excel_import.py.

Mapping:
- RS Components "RS-voorraadnr. | Beschrijving" (max 80 chars) -> Excel "Product name"
- RS Components "Beschrijving" -> Excel "Description"
- RS Components "Aantal" -> Excel "Quantity"
- RS Components "Prijs per stuk" -> Excel "Unit price"
- RS Components "RS-voorraadnr." -> Excel "Supplier Part Number"
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def find_column(fieldnames: List[str], exact_name: str, prefix_name: str) -> str:
    if exact_name in fieldnames:
        return exact_name
    for name in fieldnames:
        if name.startswith(prefix_name):
            return name
    raise ValueError(f"CSV missing required column: '{exact_name}'")


def normalize_decimal_text(value: object, field_name: str) -> str:
    text = clean_text(value)
    text = re.sub(r"[^0-9,.\-]", "", text).replace(",", ".")
    if not text:
        raise ValueError(f"Missing value for '{field_name}'.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {field_name} value '{value}'.") from exc
    if number <= 0:
        raise ValueError(f"{field_name} must be > 0 (got '{value}').")
    return format(number.normalize(), "f")


def build_product_name(stock_number: str, description: str, max_len: int = 80) -> str:
    return f"{stock_number} | {description}"[:max_len]


def parse_rs_components_csv(csv_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames or []
        if not fieldnames:
            raise ValueError("CSV has no header row.")

        stock_col = find_column(fieldnames, "RS-voorraadnr.", "RS-voorraadnr.")
        description_col = find_column(fieldnames, "Beschrijving", "Beschrijving")
        quantity_col = find_column(fieldnames, "Aantal", "Aantal")
        unit_price_col = find_column(fieldnames, "Prijs per stuk", "Prijs per stuk")

        for row in reader:
            stock_number = clean_text(row.get(stock_col, ""))
            description = clean_text(row.get(description_col, ""))
            quantity_raw = clean_text(row.get(quantity_col, ""))
            unit_price_raw = clean_text(row.get(unit_price_col, ""))

            if not stock_number and not description and not quantity_raw and not unit_price_raw:
                continue

            if not stock_number:
                raise ValueError("Missing RS-voorraadnr. for one or more rows.")

            quantity = normalize_decimal_text(quantity_raw, "Aantal")
            unit_price = normalize_decimal_text(unit_price_raw, "Prijs per stuk")
            if not description:
                description = stock_number

            rows.append(
                {
                    "product_name": build_product_name(stock_number, description),
                    "description": description,
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "supplier_part_number": stock_number,
                }
            )

    if not rows:
        raise ValueError("No item rows found in RS Components CSV.")
    return rows


def write_output_excel(output_path: Path, supplier_name: str, items: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["A2"] = supplier_name
    ws["A4"] = "Product name"
    ws["B4"] = "Description"
    ws["C4"] = "Quantity"
    ws["D4"] = "Unit price"
    ws["E4"] = "Supplier Part Number"

    row_idx = 5
    for item in items:
        ws.cell(row=row_idx, column=1, value=item["product_name"])
        ws.cell(row=row_idx, column=2, value=item["description"])
        ws.cell(row=row_idx, column=3, value=item["quantity"])
        ws.cell(row=row_idx, column=4, value=item["unit_price"])
        ws.cell(row=row_idx, column=5, value=item["supplier_part_number"])
        row_idx += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert an RS Components cart CSV into Ariba import Excel format."
    )
    parser.add_argument("--csv", required=True, help="Path to RS Components cart CSV.")
    parser.add_argument("--out", required=True, help="Path to output .xlsx file.")
    parser.add_argument(
        "--supplier",
        default="RS Components",
        help="Supplier name written to cell A2 (default: RS Components).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.csv).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()

    if not input_path.exists():
        print(f"Input CSV not found: {input_path}")
        return 2

    try:
        items = parse_rs_components_csv(input_path)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"CSV parse error: {exc}")
        return 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(output_path, args.supplier, items)
    print(f"Wrote {len(items)} item(s) to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
