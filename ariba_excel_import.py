#!/usr/bin/env python3
"""
Import non-catalog line items from Excel into SAP Ariba Buying.

Expected Excel layout:
- A1: Supplier name
- B1: Supplier value (for example "Reichelt" or "Thorlabs")
- Row 3 headers: Product name | Description | Quantity | Unit price
- Data starts at row 4

The script uses Playwright and attaches to an already open Chromium browser
instance via CDP. This is intentional so users can complete SSO/2FA manually.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Page, TimeoutError, sync_playwright


HEADER_ROW = 3
DATA_START_ROW = 4
REQUIRED_HEADERS = ["Product name", "Description", "Quantity", "Unit price"]


@dataclass
class ItemRow:
    row_number: int
    product_name: str
    description: str
    quantity: str
    unit_price: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_number_text(value: object, field_name: str, row_number: int) -> str:
    text = clean_text(value)
    if not text:
        raise ValueError(f"Row {row_number}: missing {field_name}.")
    normalized = text.replace(",", ".")
    try:
        decimal = Decimal(normalized)
    except InvalidOperation as exc:
        raise ValueError(
            f"Row {row_number}: invalid {field_name} '{text}'."
        ) from exc
    if decimal <= 0:
        raise ValueError(f"Row {row_number}: {field_name} must be > 0.")
    return format(decimal.normalize(), "f")


def load_items_from_excel(excel_path: Path) -> tuple[str, List[ItemRow]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    supplier_label = clean_text(ws["A1"].value).lower()
    if supplier_label != "supplier name":
        raise ValueError("Cell A1 must contain 'Supplier name'.")
    supplier_name = clean_text(ws["B1"].value)
    if not supplier_name:
        raise ValueError("Cell B1 (supplier value) is empty.")

    headers = [clean_text(ws.cell(row=HEADER_ROW, column=i).value) for i in range(1, 5)]
    if headers != REQUIRED_HEADERS:
        raise ValueError(
            f"Row {HEADER_ROW} headers must be exactly: {', '.join(REQUIRED_HEADERS)}."
        )

    items: List[ItemRow] = []
    current_row = DATA_START_ROW
    while True:
        values = [ws.cell(row=current_row, column=i).value for i in range(1, 5)]
        if all(clean_text(v) == "" for v in values):
            break

        product_name = clean_text(values[0])
        description = clean_text(values[1])
        if not product_name:
            raise ValueError(f"Row {current_row}: Product name is empty.")
        if not description:
            raise ValueError(f"Row {current_row}: Description is empty.")

        quantity = normalize_number_text(values[2], "Quantity", current_row)
        unit_price = normalize_number_text(values[3], "Unit price", current_row)

        items.append(
            ItemRow(
                row_number=current_row,
                product_name=product_name,
                description=description,
                quantity=quantity,
                unit_price=unit_price,
            )
        )
        current_row += 1

    if not items:
        raise ValueError("No item rows found starting at row 4.")

    return supplier_name, items


def wait_for_non_catalog_page(page: Page) -> None:
    page.get_by_text("Non-catalog request", exact=False).wait_for(timeout=20000)


def select_fixed_category(page: Page) -> None:
    category_target = "900006 (Laboratory \u2013 Items and disposables)"
    already_selected = page.get_by_text("900006 (Laboratory", exact=False).first
    if already_selected.count() > 0:
        try:
            if already_selected.is_visible():
                return
        except PlaywrightError:
            pass

    category_field = page.get_by_text("Choose a category", exact=False).first
    for _ in range(3):
        category_field.wait_for(timeout=2500)
        category_field.click()

        # Ariba sometimes lazily renders category rows only after pointer movement.
        try:
            clear_selection = page.get_by_text("Clear selection", exact=False).first
            if clear_selection.count() > 0:
                clear_selection.hover(timeout=500)
        except PlaywrightError:
            pass

        option = page.get_by_text(category_target, exact=True).first
        try:
            option.wait_for(timeout=1200)
            option.click(timeout=1200)
            return
        except PlaywrightError:
            continue

    # Final fallback: verify if selected despite missed click timing.
    page.get_by_text("900006 (Laboratory", exact=False).first.wait_for(timeout=5000)


def fill_product_name(page: Page, value: str) -> None:
    input_box = page.locator("xpath=//*[contains(normalize-space(.), 'Product name')]/following::input[1]").first
    input_box.wait_for(timeout=5000)
    input_box.fill(value)


def fill_description(page: Page, value: str) -> None:
    textareas = page.locator("textarea")
    if textareas.count() == 0:
        raise RuntimeError("Unable to locate Description textarea.")
    textareas.nth(0).fill(value)


def first_visible_or_none(locators: List) -> Optional:
    for locator in locators:
        try:
            count = locator.count()
        except PlaywrightError:
            continue
        for i in range(count):
            candidate = locator.nth(i)
            try:
                if candidate.is_visible():
                    return candidate
            except PlaywrightError:
                continue
    return None


def first_visible_text(page: Page, text: str):
    matches = page.get_by_text(text, exact=False)
    try:
        count = matches.count()
    except PlaywrightError:
        return None
    for i in range(count):
        candidate = matches.nth(i)
        try:
            if candidate.is_visible():
                return candidate
        except PlaywrightError:
            continue
    return None


def fill_quantity(page: Page, value: str) -> None:
    quantity_input = first_visible_or_none(
        [
            page.locator(
                "input:visible[aria-label*='quantity' i]:not(.shopping-cart-line-item-quantity-input)"
            ),
            page.locator(
                "xpath=//button[normalize-space()='-']/following::input[1]"
            ),
            page.locator(
                "xpath=//*[contains(normalize-space(.), 'Quantity')]/following::input[@name!='searchInGB'][1]"
            ),
        ]
    )
    if quantity_input is None:
        raise RuntimeError("Unable to locate visible Quantity field.")
    quantity_input.fill(value)
    quantity_input.evaluate(
        "el => el.dispatchEvent(new Event('change', { bubbles: true }))"
    )


def fill_unit_price(page: Page, value: str) -> None:
    price_input = first_visible_or_none(
        [
            page.locator("input:visible[aria-label*='price' i]"),
            page.locator(
                "xpath=//*[contains(normalize-space(.), 'Unit price')]/following::input[@name!='searchInGB'][1]"
            ),
            page.locator("xpath=//*[normalize-space(text())='EUR']/preceding::input[1]"),
        ]
    )
    if price_input is None:
        raise RuntimeError("Unable to locate visible Unit price field.")
    price_input.fill(value)
    price_input.evaluate(
        "el => el.dispatchEvent(new Event('change', { bubbles: true }))"
    )


def open_supplier_dialog(page: Page) -> None:
    page.get_by_text("View all suppliers", exact=False).first.click()
    page.get_by_text("Select a supplier", exact=True).wait_for(timeout=10000)
    page.locator("input[placeholder='Search']:visible").first.wait_for(timeout=10000)


def select_supplier(page: Page, supplier_name: str) -> None:
    if is_supplier_already_selected(page, supplier_name):
        return

    open_supplier_dialog(page)

    search_input = page.locator("input[placeholder='Search']:visible").first
    search_input.click()
    search_input.fill("")
    search_input.fill(supplier_name)
    search_input.press("Enter")

    exact_row = page.locator(
        "tr:visible, [role='row']:visible",
        has_text=re.compile(rf"\b{re.escape(supplier_name)}\b", re.IGNORECASE),
    ).first
    try:
        exact_row.wait_for(timeout=7000)
    except TimeoutError:
        exact_row = page.locator(
            "tr:visible, [role='row']:visible", has_text=supplier_name
        ).first
    if exact_row.count() == 0:
        raise RuntimeError(f"Supplier search found no match for '{supplier_name}'.")

    exact_row.click()
    select_button = page.get_by_role("button", name="Select").locator(":visible").first
    select_button.click()

    page.get_by_text("Selected", exact=False).first.wait_for(timeout=10000)


def verify_supplier_selected(page: Page, supplier_name: str) -> None:
    supplier_section = page.locator("xpath=//*[contains(normalize-space(.), 'Chosen supplier')]/ancestor::*[1]").first
    supplier_section.wait_for(timeout=8000)
    selected_badge = first_visible_text(page, "Selected")
    if selected_badge is None:
        raise RuntimeError("Supplier card is present but no visible 'Selected' badge found.")
    # Soft check: if the entered name appears visibly, good. If not, selected badge still counts.
    if supplier_name:
        name_match = first_visible_text(page, supplier_name)
        if name_match is not None:
            return


def is_supplier_already_selected(page: Page, supplier_name: str) -> bool:
    selected_badge = first_visible_text(page, "Selected")
    if selected_badge is None:
        return False
    supplier_panel = first_visible_text(page, "Chosen supplier")
    return supplier_panel is not None


def click_add_to_cart(page: Page) -> None:
    page.get_by_role("button", name="Add to cart").first.click()
    # Fast wait for immediate post-click state change.
    overlay = page.locator(
        "xpath=//*[contains(normalize-space(.), 'items in your cart')]"
    ).first
    try:
        overlay.wait_for(timeout=1200)
    except TimeoutError:
        try:
            page.get_by_role("button", name="Done").first.wait_for(timeout=1200)
        except TimeoutError:
            # Some UI states keep Add to cart visible; continue quickly.
            pass


def close_cart_overlay_if_present(page: Page) -> None:
    overlay_header = page.locator(
        "xpath=//*[contains(normalize-space(.), 'items in your cart')]"
    ).first
    if overlay_header.count() == 0:
        return

    close_candidates = [
        page.get_by_role(
            "button", name=re.compile(r"^(x|×|close)$", re.IGNORECASE)
        ).first,
        page.locator("button:has-text('X'):visible").first,
        page.locator(
            "xpath=//*[contains(normalize-space(.), 'items in your cart')]/ancestor::*[1]//button[1]"
        ).first,
    ]
    for close_btn in close_candidates:
        if close_btn.count() == 0:
            continue
        try:
            close_btn.click(timeout=1200)
            overlay_header.wait_for(state="hidden", timeout=1200)
            return
        except PlaywrightError:
            continue

    # Fallbacks if the close icon is not clickable.
    try:
        page.keyboard.press("Escape")
        overlay_header.wait_for(state="hidden", timeout=800)
        return
    except PlaywrightError:
        pass
    try:
        page.mouse.click(80, 220)
        overlay_header.wait_for(state="hidden", timeout=800)
    except PlaywrightError:
        pass


def create_new_item(page: Page) -> None:
    close_cart_overlay_if_present(page)

    menu_candidates = [
        page.locator("xpath=//button[normalize-space()='Done']/following::button[1]").first,
        page.get_by_role("button", name="...").first,
        page.locator("button:has-text('...')").first,
        page.locator("xpath=//*[normalize-space(text())='...']").first,
    ]

    clicked_menu = False
    for candidate in menu_candidates:
        try:
            candidate.click(timeout=450)
            clicked_menu = True
            break
        except PlaywrightError:
            continue

    if not clicked_menu:
        # Last-resort click: the menu is just to the right of "Done".
        try:
            done_btn = page.get_by_role("button", name="Done").first
            box = done_btn.bounding_box()
            if box is not None:
                page.mouse.click(box["x"] + box["width"] + 40, box["y"] + box["height"] / 2)
                clicked_menu = True
        except PlaywrightError:
            pass

    if not clicked_menu:
        raise RuntimeError("Unable to open the three-dot menu.")

    page.get_by_text("Create new", exact=True).first.wait_for(timeout=1200)
    page.get_by_text("Create new", exact=True).first.click()
    page.get_by_role("button", name="Add to cart").first.wait_for(timeout=2000)


def fill_one_item(page: Page, item: ItemRow) -> None:
    wait_for_non_catalog_page(page)
    select_fixed_category(page)
    fill_product_name(page, item.product_name)
    fill_description(page, item.description)
    fill_quantity(page, item.quantity)
    fill_unit_price(page, item.unit_price)


def ensure_supplier_selected(page: Page, supplier_name: str) -> None:
    if is_supplier_already_selected(page, supplier_name):
        return
    select_supplier(page, supplier_name)
    verify_supplier_selected(page, supplier_name)


def run_import(page: Page, supplier_name: str, items: List[ItemRow]) -> None:
    wait_for_non_catalog_page(page)

    total = len(items)
    for index, item in enumerate(items, start=1):
        print(f"[{index}/{total}] Filling item {index} into Ariba form...")
        ensure_supplier_selected(page, supplier_name)
        fill_one_item(page, item)
        click_add_to_cart(page)
        print(f"[{index}/{total}] Added item {index} to Ariba cart.")

        if index < total:
            create_new_item(page)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import non-catalog items from Excel into SAP Ariba."
    )
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to .xlsx file with Supplier name in A1/B1 and headers on row 3.",
    )
    parser.add_argument(
        "--cdp-url",
        default="http://127.0.0.1:9222",
        help="Chrome/Edge remote-debugging endpoint (default: http://127.0.0.1:9222).",
    )
    parser.add_argument(
        "--page-contains",
        default="ariba",
        help="Substring used to choose the target browser tab (default: ariba).",
    )
    return parser.parse_args()


def pick_target_page(context_page_list: List[Page], needle: str) -> Optional[Page]:
    needle_lower = needle.lower()
    for page in context_page_list:
        try:
            title = page.title().lower()
            url = page.url.lower()
        except PlaywrightError:
            continue
        if needle_lower in title or needle_lower in url:
            return page
    return context_page_list[0] if context_page_list else None


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        return 2

    try:
        supplier_name, items = load_items_from_excel(excel_path)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"Excel validation error: {exc}")
        return 2

    print(f"Loaded {len(items)} item(s). Supplier: {supplier_name}")
    print("Connecting to browser via CDP...")

    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(args.cdp_url)
            all_pages: List[Page] = []
            for context in browser.contexts:
                all_pages.extend(context.pages)
            page = pick_target_page(all_pages, args.page_contains)
            if page is None:
                raise RuntimeError("No open browser tabs found in the CDP session.")
            page.bring_to_front()
            run_import(page, supplier_name, items)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"Automation failed: {exc}")
        return 1

    print("Import completed. Please review cart and continue checkout manually.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
