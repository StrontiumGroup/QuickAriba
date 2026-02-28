#!/usr/bin/env python3
"""
Convert a DigiKey cart Excel file into the format expected by ariba_excel_import.py.

Mapping:
- DigiKey "Part Number" -> Excel "Product name"
- DigiKey "Description" -> Excel "Description"
- DigiKey "Quantity"    -> Excel "Quantity"
- DigiKey "Unit Price"  -> Excel "Unit price"

Ignored columns:
- "Manufacturer Part Number"
- "Available"
- "Backorder"
"""

from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook


HEADER_ROW = 2


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_decimal_text(value: object, field_name: str) -> str:
    text = clean_text(value)
    text = re.sub(r"[^0-9,.\-]", "", text).replace(",", ".")
    if not text:
        raise ValueError(f"Missing value for '{field_name}'.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {field_name} value '{value}'.") from exc
    if number <= 0:
        raise ValueError(f"{field_name} must be > 0 (got '{value}').")
    return format(number.normalize(), "f")


def load_digikey_rows(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = {
        clean_text(ws.cell(row=HEADER_ROW, column=col).value): col
        for col in range(1, ws.max_column + 1)
    }

    required = ["Quantity", "Part Number", "Description", "Unit Price"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing required DigiKey columns: {', '.join(missing)}")

    qty_col = headers["Quantity"]
    part_col = headers["Part Number"]
    desc_col = headers["Description"]
    price_col = headers["Unit Price"]

    rows: List[Dict[str, str]] = []
    row_idx = HEADER_ROW + 1
    while row_idx <= ws.max_row:
        part_number = clean_text(ws.cell(row=row_idx, column=part_col).value)
        description = clean_text(ws.cell(row=row_idx, column=desc_col).value)
        quantity_raw = ws.cell(row=row_idx, column=qty_col).value
        unit_price_raw = ws.cell(row=row_idx, column=price_col).value

        if (
            part_number == ""
            and description == ""
            and clean_text(quantity_raw) == ""
            and clean_text(unit_price_raw) == ""
        ):
            row_idx += 1
            continue

        if not part_number:
            # Skip non-item lines if they exist.
            row_idx += 1
            continue

        quantity = normalize_decimal_text(quantity_raw, "Quantity")
        unit_price = normalize_decimal_text(unit_price_raw, "Unit Price")
        if not description:
            description = part_number

        rows.append(
            {
                "product_name": part_number,
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
            }
        )
        row_idx += 1

    if not rows:
        raise ValueError("No item rows found in DigiKey workbook.")
    return rows


def write_output_excel(output_path: Path, supplier_name: str, items: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Supplier name"
    ws["B1"] = supplier_name
    ws["A3"] = "Product name"
    ws["B3"] = "Description"
    ws["C3"] = "Quantity"
    ws["D3"] = "Unit price"

    r = 4
    for item in items:
        ws.cell(row=r, column=1, value=item["product_name"])
        ws.cell(row=r, column=2, value=item["description"])
        ws.cell(row=r, column=3, value=item["quantity"])
        ws.cell(row=r, column=4, value=item["unit_price"])
        r += 1

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert DigiKey cart Excel to Ariba import Excel format."
    )
    parser.add_argument("--xlsx", required=True, help="Path to DigiKey cart .xlsx file.")
    parser.add_argument("--out", required=True, help="Path to output .xlsx file.")
    parser.add_argument(
        "--supplier",
        default="DigiKey",
        help="Supplier name written to cell B1 (default: DigiKey).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.xlsx).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()

    if not input_path.exists():
        print(f"Input XLSX not found: {input_path}")
        return 2

    try:
        items = load_digikey_rows(input_path)
    except Exception as exc:  # pylint: disable=broad-except
        print(f"Workbook parse error: {exc}")
        return 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(output_path, args.supplier, items)
    print(f"Wrote {len(items)} item(s) to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
